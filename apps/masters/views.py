from decimal import Decimal
from io import BytesIO

from django.contrib import messages
from django.contrib.auth.mixins import LoginRequiredMixin, PermissionRequiredMixin
from django.core.paginator import Paginator
from django.db.models import Max, Q, Sum
from django.http import Http404, HttpResponse
from django.shortcuts import redirect
from django.urls import reverse, reverse_lazy
from django.utils import timezone
from django.views import View
from django.views.generic import CreateView, DeleteView, ListView, UpdateView
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from apps.auditlog.services import log_activity, serialize_instance
from apps.common.views import EventScopedCreateView, EventScopedDeleteView, EventScopedListView, EventScopedUpdateView
from apps.inventory.models import InventoryBalance, InventoryTransaction, InventoryTransactionType, PurchaseLot
from apps.masters.forms import EventCreateForm, EventManagerContactForm, EventUpdateForm, ItemForm, SponsorForm, UpashrayForm, VendorForm, VolunteerForm
from apps.requirements.models import RequirementHeader, RequirementLine, RequirementStatus
from apps.masters.models import Event, EventManagerContact, Item, Sponsor, Upashray, Vendor, Volunteer


class EventContextMixin:
    def get_event(self):
        event_pk = self.kwargs.get("event_pk")
        if event_pk:
            return Event.objects.filter(pk=event_pk, is_active=True).first()
        return None


class EventListView(EventScopedListView):
    model = Event
    template_name = "masters/event_list.html"
    row_fields = ("name", "slug", "start_date", "end_date", "primary_contact_name", "primary_contact_mobile", "get_status_display", "is_current", "is_active")
    headers = ["Name", "Slug", "Start", "End", "Primary Contact", "Mobile", "Status", "Current", "Active"]
    create_url_name = "masters:event-create"
    edit_url_name = "masters:event-update"
    delete_url_name = "masters:event-delete"

    def get_row_url_kwargs(self, obj):
        return {"pk": obj.pk}

    def get_table_rows(self):
        rows = super().get_table_rows()
        for row in rows:
            obj = row["object"]
            row["contacts_url"] = reverse("masters:event-contact-list", kwargs={"event_pk": obj.pk})
        return rows

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["page_title"] = "Events"
        context["page_subtitle"] = "Manage Chaturmas cycles and their manager contacts."
        context["create_url"] = reverse_lazy(self.create_url_name)
        return context


class EventCreateView(LoginRequiredMixin, PermissionRequiredMixin, CreateView):
    model = Event
    form_class = EventCreateForm
    template_name = "common/form.html"
    permission_required = "masters.add_event"
    raise_exception = True

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["page_title"] = "Create Event"
        context["page_subtitle"] = "Start a new cycle with the event name and date range."
        context["list_url"] = reverse_lazy("masters:event-list")
        return context

    def form_valid(self, form):
        obj = form.save()
        Event.objects.exclude(pk=obj.pk).update(is_current=False)
        if not obj.is_current:
            obj.is_current = True
            obj.save(update_fields=["is_current"])
        self.object = obj
        log_activity(
            user=self.request.user,
            event=obj,
            action="created",
            module=self.model._meta.label_lower,
            record_id=obj.pk,
            new_value=serialize_instance(obj),
            request=self.request,
        )
        messages.success(self.request, "Record created successfully.")
        return redirect(reverse("masters:event-update", kwargs={"pk": obj.pk}))


class EventUpdateView(LoginRequiredMixin, PermissionRequiredMixin, UpdateView):
    model = Event
    form_class = EventUpdateForm
    template_name = "common/form.html"
    success_url = reverse_lazy("masters:event-list")
    permission_required = "masters.change_event"
    raise_exception = True

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["page_title"] = "Update Event"
        context["list_url"] = self.success_url
        context["manager_contacts"] = self.object.masters_eventmanagercontact_records.filter(is_active=True).order_by("-is_primary", "contact_name")
        context["manager_contacts_create_url"] = reverse("masters:event-contact-create", kwargs={"event_pk": self.object.pk})
        context["manager_contacts_list_url"] = reverse("masters:event-contact-list", kwargs={"event_pk": self.object.pk})
        context["public_share_url"] = reverse("requirements:public-collect", kwargs={"event_token": self.object.public_form_token})
        context["public_share_label"] = "Public Requirement Form Link"
        return context

    def form_valid(self, form):
        before = serialize_instance(self.get_object())
        obj = form.save()
        if obj.is_current:
            Event.objects.exclude(pk=obj.pk).update(is_current=False)
        self.object = obj
        log_activity(
            user=self.request.user,
            event=obj,
            action="updated",
            module=self.model._meta.label_lower,
            record_id=obj.pk,
            old_value=before,
            new_value=serialize_instance(obj),
            request=self.request,
        )
        messages.success(self.request, "Record updated successfully.")
        return redirect(self.success_url)


class EventDeleteView(LoginRequiredMixin, PermissionRequiredMixin, DeleteView):
    model = Event
    template_name = "common/confirm_delete.html"
    success_url = reverse_lazy("masters:event-list")
    permission_required = "masters.delete_event"
    raise_exception = True

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["list_url"] = self.success_url
        return context

    def delete(self, request, *args, **kwargs):
        obj = self.get_object()
        old_value = serialize_instance(obj)
        log_activity(
            user=request.user,
            event=obj,
            action="deleted",
            module=self.model._meta.label_lower,
            record_id=obj.pk,
            old_value=old_value,
            request=request,
        )
        messages.success(self.request, "Record deleted successfully.")
        return super().delete(request, *args, **kwargs)


class EventManagerContactListView(EventContextMixin, LoginRequiredMixin, PermissionRequiredMixin, ListView):
    model = EventManagerContact
    template_name = "common/list.html"
    row_fields = ("contact_name", "mobile", "email", "designation", "primary_label")
    headers = ["Name", "Mobile", "Email", "Designation", "Primary"]
    permission_required = "masters.view_eventmanagercontact"
    raise_exception = True
    create_url_name = "masters:event-contact-create"
    edit_url_name = "masters:event-contact-update"
    delete_url_name = "masters:event-contact-delete"

    def get_queryset(self):
        event = self.get_event()
        if event is None:
            return EventManagerContact.objects.none()
        return EventManagerContact.objects.filter(event=event, is_active=True).order_by("-is_primary", "contact_name")

    def get_row_url_kwargs(self, obj):
        return {"event_pk": obj.event_id, "pk": obj.pk}

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        event = self.get_event()
        context["page_title"] = f"{event.name} Contacts" if event else "Event Contacts"
        context["page_subtitle"] = "Manage the default manager contact and any additional contacts."
        context["create_url"] = reverse("masters:event-contact-create", kwargs={"event_pk": event.pk}) if event else ""
        context["list_url"] = reverse_lazy("masters:event-list")
        context["event"] = event
        return context


class EventManagerContactCreateView(EventContextMixin, LoginRequiredMixin, PermissionRequiredMixin, CreateView):
    model = EventManagerContact
    form_class = EventManagerContactForm
    template_name = "common/form.html"
    permission_required = "masters.add_eventmanagercontact"
    raise_exception = True

    def get_event(self):
        return super().get_event() or Event.objects.filter(is_current=True, is_active=True).first()

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        event = self.get_event()
        context["page_title"] = "Add Event Contact"
        context["page_subtitle"] = "Add a manager contact for the selected event."
        context["list_url"] = reverse("masters:event-contact-list", kwargs={"event_pk": event.pk}) if event else reverse_lazy("masters:event-list")
        context["event"] = event
        return context

    def form_valid(self, form):
        event = self.get_event()
        if event is None:
            return redirect(reverse_lazy("masters:event-list"))
        obj = form.save(commit=False)
        obj.event = event
        if obj.is_primary:
            EventManagerContact.objects.filter(event=event, is_primary=True).update(is_primary=False)
        obj.save()
        self.object = obj
        log_activity(
            user=self.request.user,
            event=event,
            action="created",
            module=self.model._meta.label_lower,
            record_id=obj.pk,
            new_value=serialize_instance(obj),
            request=self.request,
        )
        messages.success(self.request, "Record created successfully.")
        return redirect(reverse("masters:event-contact-list", kwargs={"event_pk": event.pk}))


class EventManagerContactUpdateView(EventContextMixin, LoginRequiredMixin, PermissionRequiredMixin, UpdateView):
    model = EventManagerContact
    form_class = EventManagerContactForm
    template_name = "common/form.html"
    permission_required = "masters.change_eventmanagercontact"
    raise_exception = True

    def get_queryset(self):
        event_pk = self.kwargs.get("event_pk")
        return EventManagerContact.objects.filter(event_id=event_pk).select_related("event")

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        event = self.object.event
        context["page_title"] = "Update Event Contact"
        context["page_subtitle"] = event.name if event else None
        context["list_url"] = reverse("masters:event-contact-list", kwargs={"event_pk": event.pk}) if event else reverse_lazy("masters:event-list")
        context["event"] = event
        return context

    def form_valid(self, form):
        event = form.instance.event
        before = serialize_instance(self.get_object())
        obj = form.save(commit=False)
        if obj.is_primary:
            EventManagerContact.objects.filter(event=event, is_primary=True).exclude(pk=obj.pk).update(is_primary=False)
        obj.save()
        self.object = obj
        log_activity(
            user=self.request.user,
            event=event,
            action="updated",
            module=self.model._meta.label_lower,
            record_id=obj.pk,
            old_value=before,
            new_value=serialize_instance(obj),
            request=self.request,
        )
        messages.success(self.request, "Record updated successfully.")
        return redirect(reverse("masters:event-contact-list", kwargs={"event_pk": event.pk}))


class EventManagerContactDeleteView(EventContextMixin, LoginRequiredMixin, PermissionRequiredMixin, DeleteView):
    model = EventManagerContact
    template_name = "common/confirm_delete.html"
    permission_required = "masters.delete_eventmanagercontact"
    raise_exception = True

    def get_queryset(self):
        event_pk = self.kwargs.get("event_pk")
        return EventManagerContact.objects.filter(event_id=event_pk).select_related("event")

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        event = self.object.event
        context["list_url"] = reverse("masters:event-contact-list", kwargs={"event_pk": event.pk}) if event else reverse_lazy("masters:event-list")
        return context

    def delete(self, request, *args, **kwargs):
        obj = self.get_object()
        old_value = serialize_instance(obj)
        event = obj.event
        messages.success(self.request, "Record deleted successfully.")
        response = super().delete(request, *args, **kwargs)
        log_activity(
            user=request.user,
            event=event,
            action="deleted",
            module=self.model._meta.label_lower,
            record_id=obj.pk,
            old_value=old_value,
            request=request,
        )
        return response


class ItemListView(LoginRequiredMixin, PermissionRequiredMixin, ListView):
    model = Item
    template_name = "masters/item_list.html"
    paginate_by = 50
    permission_required = "masters.view_item"
    raise_exception = True
    create_url_name = "masters:item-create"
    edit_url_name = "masters:item-update"
    delete_url_name = "masters:item-delete"

    def _perm(self, action):
        meta = self.model._meta
        return f"{meta.app_label}.{action}_{meta.model_name}"

    def get_queryset(self):
        qs = Item.objects.filter(event=self._get_event(), is_active=True)
        search = self.request.GET.get("q")
        if search:
            qs = qs.filter(
                Q(item_code__icontains=search) | Q(item_name__icontains=search) | Q(item_name_gu__icontains=search)
            )
        return qs.select_related("parent_item")

    def _get_event(self):
        event_id = self.request.GET.get("event")
        if event_id:
            return Event.objects.filter(pk=event_id, is_active=True).first()
        return Event.objects.filter(is_current=True, is_active=True).first()

    def _expand_items(self, event):
        base_items = (
            Item.objects.filter(event=event, is_active=True, parent_item__isnull=True)
            .prefetch_related("variants")
            .order_by("standard_serial", "pk")
        )
        items = []
        for base in base_items:
            variants = list(base.variants.filter(is_active=True).order_by("item_code", "pk"))
            if variants:
                items.extend(variants)
            else:
                items.append(base)
        return items

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        event = self._get_event()
        all_items = self._expand_items(event)
        search = self.request.GET.get("q", "").lower()
        if search:
            all_items = [i for i in all_items if search in i.item_code.lower() or search in i.item_name.lower() or search in i.item_name_gu.lower()]

        item_ids = [i.pk for i in all_items]
        balances = {
            b.item_id: b
            for b in InventoryBalance.objects.filter(event=event, item_id__in=item_ids)
        }
        pos_types = [InventoryTransactionType.PURCHASE, InventoryTransactionType.DONATION, InventoryTransactionType.SPONSORSHIP_RECEIPT, InventoryTransactionType.RETURN, InventoryTransactionType.ADJUSTMENT]
        acquired_qs = InventoryTransaction.objects.filter(event=event, item_id__in=item_ids, transaction_type__in=pos_types).values("item_id").annotate(total=Sum("qty"))
        acquired_map = {a["item_id"]: a["total"] for a in acquired_qs}
        packed_statuses = [RequirementStatus.PACKED, RequirementStatus.IN_PROGRESS]
        delivered_statuses = [RequirementStatus.DELIVERED, RequirementStatus.CLOSED, RequirementStatus.RECEIVED_BY_MS]
        packed_qs = RequirementLine.objects.filter(event=event, item_id__in=item_ids, requirement__status__in=packed_statuses).values("item_id").annotate(total=Sum("required_qty"))
        packed_map = {p["item_id"]: p["total"] for p in packed_qs}
        delivered_qs = RequirementLine.objects.filter(event=event, item_id__in=item_ids, requirement__status__in=delivered_statuses).values("item_id").annotate(total=Sum("required_qty"))
        delivered_map = {d["item_id"]: d["total"] for d in delivered_qs}
        req_header_ids = RequirementHeader.objects.filter(event=event, is_active=True, status__in=[RequirementStatus.CONFIRMED, RequirementStatus.NOT_CONFIRMED, RequirementStatus.SUBMITTED]).values_list("pk", flat=True)
        current_req_qs = RequirementLine.objects.filter(event=event, requirement_id__in=req_header_ids, item_id__in=item_ids).values("item_id").annotate(total=Sum("required_qty"))
        current_req_map = {r["item_id"]: r["total"] for r in current_req_qs}
        latest_lots = {}
        for item_id in item_ids:
            lot = PurchaseLot.objects.filter(event=event, item_id=item_id).order_by("-transaction_date", "-created_at").first()
            if lot:
                latest_lots[item_id] = lot

        table_rows = []
        for item in all_items:
            bal = balances.get(item.pk)
            qty_acquired = int(acquired_map.get(item.pk, 0)) if acquired_map.get(item.pk, 0) == int(acquired_map.get(item.pk, 0)) else acquired_map.get(item.pk, 0)
            qty_packed = int(packed_map.get(item.pk, 0)) if packed_map.get(item.pk, 0) == int(packed_map.get(item.pk, 0)) else packed_map.get(item.pk, 0)
            qty_delivered = int(delivered_map.get(item.pk, 0)) if delivered_map.get(item.pk, 0) == int(delivered_map.get(item.pk, 0)) else delivered_map.get(item.pk, 0)
            current_stock = qty_acquired - (qty_packed + qty_delivered)
            lot = latest_lots.get(item.pk)
            cost = lot.unit_rate if lot else Decimal(item.estimated_rate or 0)
            if cost == int(cost):
                cost = int(cost)
            vendor = str(lot.vendor) if lot and lot.vendor else ""
            manager = lot.managed_by.get_full_name() if lot and lot.managed_by else (str(lot.managed_by) if lot and lot.managed_by else "")
            table_rows.append({
                "item_code": item.item_code,
                "display_name": item.display_name(),
                "type_size": item.variant_name_gu or item.variant_name or item.default_size_gu or item.default_size or "",
                "current_req": int(current_req_map.get(item.pk, 0)),
                "current_stock": current_stock,
                "qty_acquired": qty_acquired,
                "qty_packed": qty_packed,
                "qty_delivered": qty_delivered,
                "cost": cost,
                "vendor": vendor,
                "manager": manager,
                "edit_url": reverse(self.edit_url_name, kwargs={"pk": item.pk}) if self.edit_url_name else "",
                "delete_url": reverse(self.delete_url_name, kwargs={"pk": item.pk}) if self.delete_url_name else "",
            })

        paginator = Paginator(table_rows, self.paginate_by)
        page_number = self.request.GET.get("page", 1)
        page_obj = paginator.get_page(page_number)

        context["page_title"] = "Items"
        context["page_subtitle"] = "Event-scoped item master"
        context["create_url"] = reverse_lazy(self.create_url_name)
        context["table_rows"] = page_obj.object_list
        context["page_obj"] = page_obj
        context["paginator"] = paginator
        context["event_queryset"] = Event.objects.filter(is_active=True).order_by("-is_current", "-start_date", "name")
        context["can_add"] = self.request.user.has_perm(self._perm("add"))
        context["can_change"] = self.request.user.has_perm(self._perm("change"))
        context["can_delete"] = self.request.user.has_perm(self._perm("delete"))
        reqs = RequirementHeader.objects.filter(event=event, is_active=True).exclude(status=RequirementStatus.DRAFT)
        context["order_summary"] = {
            "confirmed": reqs.filter(status__in=[RequirementStatus.CONFIRMED, RequirementStatus.NOT_CONFIRMED]).count(),
            "packed": reqs.filter(status__in=[RequirementStatus.PACKED, RequirementStatus.IN_PROGRESS]).count(),
            "delivered": reqs.filter(status__in=[RequirementStatus.DELIVERED, RequirementStatus.CLOSED, RequirementStatus.RECEIVED_BY_MS]).count(),
            "total": reqs.count(),
        }
        return context


class ItemListExportView(LoginRequiredMixin, View):
    def get(self, request, *args, **kwargs):
        event_id = request.GET.get("event")
        event = Event.objects.filter(pk=event_id, is_active=True).first() if event_id else Event.objects.filter(is_current=True, is_active=True).first()
        if event is None:
            return HttpResponse("No active event found.", status=404)

        base_items = Item.objects.filter(event=event, is_active=True, parent_item__isnull=True).prefetch_related("variants").order_by("standard_serial", "pk")
        all_items = []
        for base in base_items:
            variants = list(base.variants.filter(is_active=True).order_by("item_code", "pk"))
            if variants:
                all_items.extend(variants)
            else:
                all_items.append(base)

        item_ids = [i.pk for i in all_items]
        balances = {b.item_id: b for b in InventoryBalance.objects.filter(event=event, item_id__in=item_ids)}
        pos_types = [InventoryTransactionType.PURCHASE, InventoryTransactionType.DONATION, InventoryTransactionType.SPONSORSHIP_RECEIPT, InventoryTransactionType.RETURN, InventoryTransactionType.ADJUSTMENT]
        acquired_map = {a["item_id"]: a["total"] for a in InventoryTransaction.objects.filter(event=event, item_id__in=item_ids, transaction_type__in=pos_types).values("item_id").annotate(total=Sum("qty"))}
        packed_statuses = [RequirementStatus.PACKED, RequirementStatus.IN_PROGRESS]
        delivered_statuses = [RequirementStatus.DELIVERED, RequirementStatus.CLOSED, RequirementStatus.RECEIVED_BY_MS]
        packed_map = {p["item_id"]: p["total"] for p in RequirementLine.objects.filter(event=event, item_id__in=item_ids, requirement__status__in=packed_statuses).values("item_id").annotate(total=Sum("required_qty"))}
        delivered_map = {d["item_id"]: d["total"] for d in RequirementLine.objects.filter(event=event, item_id__in=item_ids, requirement__status__in=delivered_statuses).values("item_id").annotate(total=Sum("required_qty"))}
        req_header_ids = RequirementHeader.objects.filter(event=event, is_active=True, status__in=[RequirementStatus.CONFIRMED, RequirementStatus.NOT_CONFIRMED, RequirementStatus.SUBMITTED]).values_list("pk", flat=True)
        current_req_map = {r["item_id"]: r["total"] for r in RequirementLine.objects.filter(event=event, requirement_id__in=req_header_ids, item_id__in=item_ids).values("item_id").annotate(total=Sum("required_qty"))}
        stock_map = {}
        for item_id in item_ids:
            acquired = acquired_map.get(item_id, 0) or 0
            packed = packed_map.get(item_id, 0) or 0
            delivered = delivered_map.get(item_id, 0) or 0
            stock_map[item_id] = int(acquired - packed - delivered) if (acquired - packed - delivered) == int(acquired - packed - delivered) else (acquired - packed - delivered)
        latest_lots = {}
        for item_id in item_ids:
            lot = PurchaseLot.objects.filter(event=event, item_id=item_id).order_by("-transaction_date", "-created_at").first()
            if lot:
                latest_lots[item_id] = lot

        workbook = Workbook()

        center = Alignment(horizontal="center", vertical="center")
        right = Alignment(horizontal="right", vertical="center")
        header_fill = PatternFill("solid", fgColor="DCE9F5")

        # --- Sheet 1: Order Summary (pivot) ---
        ws_summary = workbook.active
        ws_summary.title = "Order Summary"

        summary_headers = ["Item Code", "Item Name / Variant", "Type / Size", "Current Req.", "Current Stock", "Qty Acquired", "Qty Packed", "Qty Delivered"]
        for col, h in enumerate(summary_headers, 1):
            cell = ws_summary.cell(row=1, column=col, value=h)
            cell.fill = header_fill
            cell.font = Font(bold=True)
            cell.alignment = center

        for idx, item in enumerate(all_items, 2):
            ws_summary.cell(row=idx, column=1, value=item.item_code)
            ws_summary.cell(row=idx, column=2, value=item.display_name())
            ws_summary.cell(row=idx, column=3, value=item.variant_name_gu or item.variant_name or item.default_size_gu or item.default_size or "")
            ws_summary.cell(row=idx, column=4, value=int(current_req_map.get(item.pk, 0)))
            ws_summary.cell(row=idx, column=5, value=stock_map.get(item.pk, 0))
            ws_summary.cell(row=idx, column=6, value=int(acquired_map.get(item.pk, 0)))
            ws_summary.cell(row=idx, column=7, value=int(packed_map.get(item.pk, 0)))
            ws_summary.cell(row=idx, column=8, value=int(delivered_map.get(item.pk, 0)))

        ws_summary.freeze_panes = "A2"
        for col_letter, w in [("A", 12), ("B", 75), ("C", 15), ("D", 12), ("E", 12), ("F", 12), ("G", 12), ("H", 12)]:
            ws_summary.column_dimensions[col_letter].width = w
        wrap_left_mid = Alignment(horizontal="left", vertical="center", wrap_text=True)
        for row in ws_summary.iter_rows(min_row=1, max_row=ws_summary.max_row):
            ws_summary.row_dimensions[row[0].row].height = 27
            for cell in row:
                cell.alignment = wrap_left_mid
        ws_summary.print_area = f"A1:H{ws_summary.max_row}"
        ws_summary.print_title_rows = "1:1"

        # --- Sheet 2: Response Sheet (one row per form) ---
        ws_response = workbook.create_sheet("Response Sheet")

        response_headers_qs = RequirementHeader.objects.filter(event=event).exclude(status=RequirementStatus.DRAFT).select_related("upashray").prefetch_related("lines__item").order_by("created_at")

        alphabet = "ABCDEFGHIJKLMNOPQRSTUVWXYZ"
        response_items = []
        item_serial_map = {}
        for base in base_items:
            variants = list(base.variants.filter(is_active=True).order_by("item_code", "pk"))
            if variants:
                for vi, variant in enumerate(variants):
                    suffix = alphabet[vi] if vi < 26 else f"X{vi+1}"
                    response_items.append(variant)
                    item_serial_map[variant.pk] = f"{base.standard_serial or base.pk}-{suffix}"
            else:
                response_items.append(base)
                item_serial_map[base.pk] = str(base.standard_serial or base.pk)

        total_fill = PatternFill("solid", fgColor="FFF3BF")
        bold_font = Font(bold=True, color="14324F")

        item_col_map = {}
        for item in response_items:
            display = item.item_name_gu or item.item_name
            item_col_map[item.pk] = {
                "col_idx": len(item_col_map) + 1,
                "header": f"{item_serial_map[item.pk]}-{display}",
            }

        basic_headers = [
            "Sr. No.", "Timestamp", "Form No.", "Order ID", "Requirement Date",
            "Route", "Sub Route", "Pujya Shri", "Thana", "Area",
            "Current Address", "Chaturmas Address", "Chaturmas Entry Date",
            "Volunteer Name", "Volunteer Mobile", "Stay Type",
            "Care Taker Name", "Care Taker Mobile", "Status",
        ]
        all_headers = basic_headers + [v["header"] for v in item_col_map.values()] + ["Total Qty", "Extra Item 1", "Extra Item 2", "Extra Item 3", "Extra Item 4"]
        for col, h in enumerate(all_headers, 1):
            cell = ws_response.cell(row=1, column=col, value=h)
            cell.fill = header_fill
            cell.font = bold_font
            cell.alignment = center

        # Row 2: Type/Size of each item column
        for col, h in enumerate(all_headers, 1):
            if col > len(basic_headers) and col <= len(basic_headers) + len(item_col_map):
                item_pk = list(item_col_map.keys())[col - len(basic_headers) - 1]
                item_obj = next((i for i in response_items if i.pk == item_pk), None)
                if item_obj:
                    ws_response.cell(row=2, column=col, value=item_obj.variant_name_gu or item_obj.variant_name or item_obj.default_size_gu or item_obj.default_size or "")
                    ws_response.cell(row=2, column=col).font = Font(italic=True, color="666666")

        DATA_START_ROW = 3

        form_count = 0
        totals = [0] * (len(item_col_map) + 1)

        for header in response_headers_qs:
            form_count += 1
            row_qty_total = 0
            item_qty_map = {}
            for line in header.lines.all():
                qty = float(line.required_qty)
                item_qty_map[line.item_id] = qty
                row_qty_total += qty

            row_data = [
                form_count,
                timezone.localtime(header.created_at).strftime("%d-%b-%Y %H:%M") if header.created_at else "",
                header.form_number or "",
                header.order_number or "",
                header.requirement_date.strftime("%d-%b-%Y") if header.requirement_date else "",
                header.get_route_area_display() or "",
                header.get_route_sub_area_display() or "",
                header.pujya_shri_name or "",
                header.thana_count or "",
                header.area or "",
                header.current_address or "",
                header.chaturmas_place_address or "",
                header.chaturmas_entry_date.strftime("%d-%b-%Y") if header.chaturmas_entry_date else "",
                header.volunteer_name or "",
                header.volunteer_mobile or "",
                header.get_stay_type_display() or "",
                header.caretaker_name or "",
                header.caretaker_mobile or "",
                header.get_status_display(),
            ]

            item_cells = []
            for item_pk, info in item_col_map.items():
                qty = item_qty_map.get(item_pk, 0)
                item_cells.append(qty)
                totals[info["col_idx"] - 1] += qty
            totals[-1] += row_qty_total

            row_data.extend(item_cells)
            row_data.append(row_qty_total)
            note_lines = (header.remarks.splitlines() if header.remarks else [])[:4]
            note_lines = (note_lines + ["", "", "", ""])[:4]
            row_data.extend(note_lines)
            for col, val in enumerate(row_data, 1):
                ws_response.cell(row=form_count + DATA_START_ROW - 1, column=col, value=val)

        total_row_data = ["TOTAL", "", f"{form_count} Forms"] + [""] * (len(basic_headers) - 3)
        for item_pk, info in item_col_map.items():
            total_row_data.append(totals[info["col_idx"] - 1])
        total_row_data.append(totals[-1])
        total_row = form_count + DATA_START_ROW
        for col, val in enumerate(total_row_data, 1):
            cell = ws_response.cell(row=total_row, column=col, value=val)
            cell.fill = total_fill
            cell.font = bold_font
            cell.alignment = center

        ws_response.freeze_panes = "A3"
        basic_count = len(basic_headers)
        item_start_col = basic_count + 1
        total_qty_col = basic_count + len(item_col_map) + 1
        extra_cols_start = total_qty_col + 1

        alt_fill_1 = PatternFill("solid", fgColor="E8F0FE")
        alt_fill_2 = PatternFill("solid", fgColor="FDF2E9")
        total_qty_fill = PatternFill("solid", fgColor="F9E79F")
        total_qty_font = Font(bold=True, color="14324F", size=11)

        total_qty_cell = ws_response.cell(row=total_row, column=total_qty_col)
        total_qty_cell.fill = total_qty_fill
        total_qty_cell.font = total_qty_font

        for col_idx in range(1, len(all_headers) + 1):
            if col_idx <= basic_count:
                max_len = 0
                for row in ws_response.iter_rows(min_row=1, max_row=ws_response.max_row, min_col=col_idx, max_col=col_idx):
                    for cell in row:
                        try:
                            text = str(cell.value or "")
                        except Exception:
                            text = ""
                        max_len = max(max_len, len(text))
                ws_response.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 40)
            elif col_idx <= total_qty_col:
                ws_response.column_dimensions[get_column_letter(col_idx)].width = 15
            else:
                ws_response.column_dimensions[get_column_letter(col_idx)].width = 15

        for row_idx in range(1, ws_response.max_row + 1):
            for col_idx in range(item_start_col, extra_cols_start + 4 + 1):
                cell = ws_response.cell(row=row_idx, column=col_idx)
                if row_idx >= DATA_START_ROW and row_idx <= ws_response.max_row - 1:
                    if col_idx == total_qty_col:
                        cell.fill = total_qty_fill
                        cell.font = total_qty_font
                    else:
                        item_offset = col_idx - item_start_col
                        cell.fill = alt_fill_1 if item_offset % 2 == 0 else alt_fill_2
                cell.alignment = center

        for row_idx in range(DATA_START_ROW, ws_response.max_row):
            for col_idx in range(1, total_qty_col + 1):
                ws_response.cell(row=row_idx, column=col_idx).alignment = center

        ws_extra = workbook.create_sheet("Extra Items")
        extra_headers = ["Form No.", "Extra Item", "Form Status"]
        for col, h in enumerate(extra_headers, 1):
            cell = ws_extra.cell(row=1, column=col, value=h)
            cell.fill = header_fill
            cell.font = Font(bold=True)
            cell.alignment = center

        extra_row = 2
        for header in response_headers_qs:
            form_no = header.form_number or ""
            status_display = header.get_status_display()
            note_lines = (header.remarks.splitlines() if header.remarks else [])[:4]
            note_lines = [n.strip() for n in note_lines if n.strip()]
            for note in note_lines:
                ws_extra.cell(row=extra_row, column=1, value=form_no)
                ws_extra.cell(row=extra_row, column=2, value=note)
                ws_extra.cell(row=extra_row, column=3, value=status_display)
                extra_row += 1

        if extra_row == 2:
            ws_extra.cell(row=2, column=1, value="(No extra items)")
            extra_row = 3

        ws_extra.column_dimensions["A"].width = 18
        ws_extra.column_dimensions["B"].width = 50
        ws_extra.column_dimensions["C"].width = 18
        ws_extra.freeze_panes = "A2"

        buffer = BytesIO()
        workbook.save(buffer)
        response = HttpResponse(buffer.getvalue(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response["Content-Disposition"] = f'attachment; filename="item_master_{event.name}.xlsx"'
        return response


class ItemCreateView(EventScopedCreateView):
    model = Item
    form_class = ItemForm
    template_name = "common/form.html"
    success_url = reverse_lazy("masters:item-list")

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs["user"] = self.request.user
        return kwargs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["page_title"] = "Create Item"
        context["list_url"] = self.success_url
        return context

    def form_valid(self, form):
        event = self.get_current_event()
        if event is None:
            raise Http404("No active event found.")
        obj = form.save(commit=False)
        obj.event = event
        next_serial = (Item.objects.filter(event=event).aggregate(max_serial=Max("standard_serial"))["max_serial"] or 0) + 1
        if not obj.standard_serial:
            obj.standard_serial = next_serial
        obj.is_active = bool(form.cleaned_data.get("add_to_current_form_immediately"))
        obj.save()
        self.object = obj
        if hasattr(form, "save_m2m"):
            form.save_m2m()
        log_activity(
            user=self.request.user,
            event=event,
            action="created",
            module=self.model._meta.label_lower,
            record_id=obj.pk,
            new_value=serialize_instance(obj),
            request=self.request,
        )
        messages.success(self.request, "Record created successfully.")
        return redirect(self.success_url)


class ItemUpdateView(EventScopedUpdateView):
    model = Item
    form_class = ItemForm
    template_name = "common/form.html"
    success_url = reverse_lazy("masters:item-list")

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs["user"] = self.request.user
        return kwargs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["page_title"] = "Update Item"
        context["list_url"] = self.success_url
        return context

    def form_valid(self, form):
        before = serialize_instance(self.get_object())
        obj = form.save(commit=False)
        obj.updated_by = self.request.user
        if "is_active" in form.cleaned_data:
            obj.is_active = form.cleaned_data["is_active"]
        obj.save()
        self.object = obj
        log_activity(
            user=self.request.user,
            event=obj.event,
            action="updated",
            module=self.model._meta.label_lower,
            record_id=obj.pk,
            old_value=before,
            new_value=serialize_instance(obj),
            request=self.request,
        )
        messages.success(self.request, "Record updated successfully.")
        return redirect(self.success_url)


class ItemDeleteView(EventScopedDeleteView):
    model = Item
    template_name = "common/confirm_delete.html"
    success_url = reverse_lazy("masters:item-list")

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["list_url"] = self.success_url
        return context


class UpashrayListView(EventScopedListView):
    model = Upashray
    template_name = "common/list.html"
    row_fields = ("name", "display_sub_area", "city", "contact_person", "mobile", "get_status_display")
    headers = ["Name", "Route", "City", "Contact", "Mobile", "Status"]
    search_fields = ["name", "area", "city", "mobile"]
    create_url_name = "masters:upashray-create"
    edit_url_name = "masters:upashray-update"
    delete_url_name = "masters:upashray-delete"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["page_title"] = "Upashrays"
        context["page_subtitle"] = "Event-wise upashray master"
        context["create_url"] = reverse_lazy(self.create_url_name)
        return context


class UpashrayCreateView(EventScopedCreateView):
    model = Upashray
    form_class = UpashrayForm
    template_name = "common/form.html"
    success_url = reverse_lazy("masters:upashray-list")

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["page_title"] = "Create Upashray"
        context["list_url"] = self.success_url
        return context


class UpashrayUpdateView(EventScopedUpdateView):
    model = Upashray
    form_class = UpashrayForm
    template_name = "common/form.html"
    success_url = reverse_lazy("masters:upashray-list")

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["page_title"] = "Update Upashray"
        context["list_url"] = self.success_url
        return context


class UpashrayDeleteView(EventScopedDeleteView):
    model = Upashray
    template_name = "common/confirm_delete.html"
    success_url = reverse_lazy("masters:upashray-list")

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["list_url"] = self.success_url
        return context


class VolunteerListView(EventScopedListView):
    model = Volunteer
    template_name = "common/list.html"
    row_fields = ("name", "mobile", "email", "area", "vehicle_available")
    headers = ["Name", "Mobile", "Email", "Area", "Vehicle"]
    search_fields = ["name", "mobile", "email", "area"]
    create_url_name = "masters:volunteer-create"
    edit_url_name = "masters:volunteer-update"
    delete_url_name = "masters:volunteer-delete"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["page_title"] = "Volunteers"
        context["create_url"] = reverse_lazy(self.create_url_name)
        return context


class VolunteerCreateView(EventScopedCreateView):
    model = Volunteer
    form_class = VolunteerForm
    template_name = "common/form.html"
    success_url = reverse_lazy("masters:volunteer-list")

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["page_title"] = "Create Volunteer"
        context["list_url"] = self.success_url
        return context


class VolunteerUpdateView(EventScopedUpdateView):
    model = Volunteer
    form_class = VolunteerForm
    template_name = "common/form.html"
    success_url = reverse_lazy("masters:volunteer-list")

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["page_title"] = "Update Volunteer"
        context["list_url"] = self.success_url
        return context


class VolunteerDeleteView(EventScopedDeleteView):
    model = Volunteer
    template_name = "common/confirm_delete.html"
    success_url = reverse_lazy("masters:volunteer-list")

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["list_url"] = self.success_url
        return context


class SponsorListView(EventScopedListView):
    model = Sponsor
    template_name = "common/list.html"
    row_fields = ("sponsor_name", "mobile", "organization", "reference_volunteer")
    headers = ["Sponsor", "Mobile", "Organization", "Reference Volunteer"]
    search_fields = ["sponsor_name", "mobile", "organization"]
    create_url_name = "masters:sponsor-create"
    edit_url_name = "masters:sponsor-update"
    delete_url_name = "masters:sponsor-delete"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["page_title"] = "Sponsors"
        context["create_url"] = reverse_lazy(self.create_url_name)
        return context


class SponsorCreateView(EventScopedCreateView):
    model = Sponsor
    form_class = SponsorForm
    template_name = "common/form.html"
    success_url = reverse_lazy("masters:sponsor-list")

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["page_title"] = "Create Sponsor"
        context["list_url"] = self.success_url
        return context


class SponsorUpdateView(EventScopedUpdateView):
    model = Sponsor
    form_class = SponsorForm
    template_name = "common/form.html"
    success_url = reverse_lazy("masters:sponsor-list")

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["page_title"] = "Update Sponsor"
        context["list_url"] = self.success_url
        return context


class SponsorDeleteView(EventScopedDeleteView):
    model = Sponsor
    template_name = "common/confirm_delete.html"
    success_url = reverse_lazy("masters:sponsor-list")

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["list_url"] = self.success_url
        return context


class VendorListView(EventScopedListView):
    model = Vendor
    template_name = "common/list.html"
    row_fields = ("vendor_name", "contact_person", "mobile", "gst_no")
    headers = ["Vendor", "Contact", "Mobile", "GST No"]
    search_fields = ["vendor_name", "contact_person", "mobile", "gst_no"]
    create_url_name = "masters:vendor-create"
    edit_url_name = "masters:vendor-update"
    delete_url_name = "masters:vendor-delete"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["page_title"] = "Vendors"
        context["create_url"] = reverse_lazy(self.create_url_name)
        return context


class VendorCreateView(EventScopedCreateView):
    model = Vendor
    form_class = VendorForm
    template_name = "common/form.html"
    success_url = reverse_lazy("masters:vendor-list")

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["page_title"] = "Create Vendor"
        context["list_url"] = self.success_url
        return context


class VendorUpdateView(EventScopedUpdateView):
    model = Vendor
    form_class = VendorForm
    template_name = "common/form.html"
    success_url = reverse_lazy("masters:vendor-list")

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["page_title"] = "Update Vendor"
        context["list_url"] = self.success_url
        return context


class VendorDeleteView(EventScopedDeleteView):
    model = Vendor
    template_name = "common/confirm_delete.html"
    success_url = reverse_lazy("masters:vendor-list")

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["list_url"] = self.success_url
        return context
