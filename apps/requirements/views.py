import json
import uuid
from collections import defaultdict
from io import BytesIO
from pathlib import Path
from xml.sax.saxutils import escape

from django.contrib import messages
from django.conf import settings
from django.contrib.auth.mixins import LoginRequiredMixin
from django.db import transaction
from django.http import HttpResponse, JsonResponse
from django.forms import formset_factory
from django.shortcuts import get_object_or_404, redirect, render
from django.template.loader import render_to_string
from django.urls import reverse, reverse_lazy

from .services import send_form_backup_email
from django.views import View
from django.utils import timezone

from reportlab.lib import colors
from reportlab.lib.enums import TA_LEFT
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.graphics.barcode import createBarcodeDrawing

from reportlab.platypus import Image, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

try:
    from weasyprint import HTML, CSS
except Exception:  # pragma: no cover - optional dependency
    HTML = None
    CSS = None

from apps.common.pdf_utils import (
    GUJARATI_FONT_NAME as PDF_UTILS_GUJARATI_FONT,
    _gujarati_font_registered as PDF_GUJARATI_FONT_REGISTERED,
    generate_gujarati_pdf_fpdf2,
    generate_weasyprint_pdf,
    NumberedCanvas as SharedNumberedCanvas,
)

from apps.common.views import EventScopedCreateView, EventScopedDeleteView, EventScopedListView, EventScopedUpdateView
from apps.dashboard.services import build_public_item_preview
from apps.inventory.services import apply_requirement_packing
from apps.masters.models import Event, EventManagerContact, Item, ItemCategory
from apps.requirements.forms import (
    RequirementCollectionHeaderForm,
    RequirementCollectionItemForm,
    RequirementHeaderForm,
    RequirementLineForm,
    SpecialRequirementForm,
    ViewControlForm,
)
from apps.accounts.permissions import is_manager
from apps.requirements.models import EditRequest, RequirementHeader, RequirementLine, RequirementStatus, SpecialRequirement, ViewControl
from apps.masters.models import Upashray


PUBLIC_STATUS_CHOICES = [
    (RequirementStatus.DRAFT, "Open"),
    (RequirementStatus.NOT_CONFIRMED, "Not Confirmed"),
    (RequirementStatus.CONFIRMED, "Confirmed"),
    (RequirementStatus.PACKED, "Packed"),
    (RequirementStatus.DELIVERED, "Delivered"),
    (RequirementStatus.SUBMITTED, "Pending"),
    (RequirementStatus.IN_PROGRESS, "Packing done"),
    (RequirementStatus.CLOSED, "On route"),
    (RequirementStatus.RECEIVED_BY_MS, "Recieved by M.S."),
    (RequirementStatus.CANCELLED, "Rejected by M.S."),
    (RequirementStatus.RETURN_REQUESTED, "Return Requested"),
    (RequirementStatus.RETURN_DONE, "Return Done"),
]


CATEGORY_LABELS = {
    "en": {
        ItemCategory.GENERAL: "General",
        ItemCategory.STATIONERY: "Stationery",
        ItemCategory.MEDICAL: "Medical",
        ItemCategory.AYURVEDIC: "Ayurvedic",
        ItemCategory.COLOR_MATERIAL: "Color Material",
    },
    "gu": {
        ItemCategory.GENERAL: "જનરલ વસ્તુઓ",
        ItemCategory.STATIONERY: "સ્ટેશનરી",
        ItemCategory.MEDICAL: "મેડિકલ",
        ItemCategory.AYURVEDIC: "આયુર્વેદિક દવાઓ",
        ItemCategory.COLOR_MATERIAL: "રંગ સામગ્રી",
    },
}

CATEGORY_ROW_CLASSES = {
    ItemCategory.GENERAL: "cat-general",
    ItemCategory.STATIONERY: "cat-stationery",
    ItemCategory.MEDICAL: "cat-medical",
    ItemCategory.AYURVEDIC: "cat-ayurvedic",
    ItemCategory.COLOR_MATERIAL: "cat-color",
}

PDF_FONT_NAME = "Helvetica"
PDF_GUJARATI_FONT_NAME = PDF_UTILS_GUJARATI_FONT
if not PDF_GUJARATI_FONT_REGISTERED:
    _fallback_candidates = [
        Path(settings.BASE_DIR) / "assets/fonts/NotoSansGujarati-Regular.ttf",
        Path(settings.GUJARATI_FONT_PATH),
        Path("C:/Windows/Fonts/shruti.ttf"),
        Path("C:/Windows/Fonts/Nirmala.ttc"),
        Path("/usr/share/fonts/truetype/noto/NotoSansGujarati-Regular.ttf"),
        Path("/usr/share/fonts/opentype/noto/NotoSansGujarati-Regular.ttf"),
        Path("/usr/share/fonts/truetype/lohit-gujarati/Lohit-Gujarati.ttf"),
    ]
    for candidate in _fallback_candidates:
        if candidate.exists():
            try:
                pdfmetrics.registerFont(TTFont("KMMGujarati", str(candidate)))
                PDF_GUJARATI_FONT_NAME = "KMMGujarati"
                break
            except Exception:
                pass


def _resolve_font_path():
    path = Path(settings.GUJARATI_FONT_PATH)
    if path.exists():
        return str(path)
    candidates = [
        Path(settings.BASE_DIR) / "static" / "fonts" / "NotoSansGujarati-Regular.ttf",
        Path(settings.BASE_DIR) / "assets/fonts/NotoSansGujarati-Regular.ttf",
    ]
    for candidate in candidates:
        if candidate.exists():
            return str(candidate)
    return settings.GUJARATI_FONT_PATH


def _lang_code(request):
    requested = (request.GET.get("lang") or request.POST.get("lang") or "").lower()
    if requested.startswith("gu"):
        return "gu"
    if requested.startswith("en"):
        return "en"
    return "gu"


def _with_lang(url, language_code):
    separator = "&" if "?" in url else "?"
    return f"{url}{separator}lang={language_code}"


def _format_qty(value):
    if value is None:
        return "-"
    try:
        numeric = float(value)
    except (TypeError, ValueError):
        return str(value)
    if numeric.is_integer():
        return str(int(numeric))
    return ("{:.3f}".format(numeric)).rstrip("0").rstrip(".")


def _format_pdf_date(value):
    if not value:
        return "-"
    month_names = [
        "January",
        "February",
        "March",
        "April",
        "May",
        "June",
        "July",
        "August",
        "September",
        "October",
        "November",
        "December",
    ]
    return f"{value.day:02d}-{month_names[value.month - 1]}-{value.year}"


def _format_main_contact(event):
    if event is None:
        return "Main Event Manager: -"
    contact = EventManagerContact.objects.filter(event=event, is_primary=True).first()
    if contact:
        bits = [contact.contact_name or "-", contact.mobile or ""]
        contact_text = " | ".join(bit for bit in bits if bit)
        return f"Main Event Manager: {contact_text}"
    bits = [event.primary_contact_name or "-", event.primary_contact_mobile or ""]
    contact_text = " | ".join(bit for bit in bits if bit)
    return f"Main Event Manager: {contact_text}"


def _item_name_for_language(item, language_code):
    if language_code == "gu":
        if item.parent_item_id:
            base = item.parent_item.item_name_gu or item.parent_item.item_name
            variant = item.variant_name_gu or item.variant_name or ""
            return f"{base} - {variant}" if variant else base
        return item.item_name_gu or item.item_name
    if item.parent_item_id:
        base = item.parent_item.item_name
        variant = item.variant_name or ""
        return f"{base} - {variant}" if variant else base
    return item.item_name


def _item_size_for_language(item, language_code):
    if language_code == "gu":
        if item.parent_item_id:
            return item.variant_name_gu or item.variant_name or item.default_size_gu or item.default_size or "-"
        return item.default_size_gu or item.default_size or "-"
    if item.parent_item_id:
        return item.variant_name or item.variant_name_gu or item.default_size or "-"
    return item.default_size or "-"


def _variant_suffix(index):
    alphabet = "ABCDEFGHIJKLMNOPQRSTUVWXYZ"
    if index < len(alphabet):
        return alphabet[index]
    return f"X{index + 1}"


def _group_requirement_lines(lines):
    grouped = []
    current_key = None
    current_group = None
    for line in sorted(lines, key=_line_sort_key):
        item = line.item
        base_item = item.parent_item if item.parent_item_id else item
        key = base_item.pk
        if key != current_key:
            current_key = key
            current_group = {
                "base_item": base_item,
                "lines": [],
            }
            grouped.append(current_group)
        if item.parent_item_id:
            siblings = list(base_item.variants.filter(is_active=True).order_by("variant_name", "pk"))
            suffix_index = next((idx for idx, sibling in enumerate(siblings) if sibling.pk == item.pk), 0)
            serial_display = f"{base_item.standard_serial or base_item.pk}-{_variant_suffix(suffix_index)}"
            size_display = item.variant_name_gu or item.variant_name or item.default_size or ""
            display_name = f"{item.item_name_gu or item.item_name} - {item.variant_name_gu or item.variant_name or ''}".strip(" -")
        else:
            serial_display = str(item.standard_serial or item.pk)
            size_display = item.default_size or ""
            display_name = item.item_name_gu or item.item_name
        current_group["lines"].append(
            {
                "line": line,
                "serial_display": serial_display,
                "display_name": display_name,
                "size_display": size_display,
            }
        )
    return grouped


def _line_serial_display(item):
    if not item.parent_item_id:
        return str(item.standard_serial or item.pk)
    base_item = item.parent_item
    siblings = list(base_item.variants.filter(is_active=True).order_by("variant_name", "pk"))
    suffix_index = next((idx for idx, sibling in enumerate(siblings) if sibling.pk == item.pk), 0)
    return f"{base_item.standard_serial or base_item.pk}-{_variant_suffix(suffix_index)}"


def _line_sort_key(line):
    item = line.item
    base_item = item.parent_item if item.parent_item_id else item
    return (base_item.standard_serial or base_item.pk, 1 if item.parent_item_id else 0, item.variant_name or "", item.pk)


def _requirement_pdf_rows(lines, language_code="gu", filter_zero=True):
    rows = []
    for line in sorted(lines, key=_line_sort_key):
        if filter_zero and (not line.required_qty or line.required_qty <= 0):
            continue
        item = line.item
        qty_display = _format_qty(line.required_qty)
        if not filter_zero and (not line.required_qty or line.required_qty <= 0):
            qty_display = "--"
        rows.append(
            (
                _line_serial_display(item),
                _item_name_for_language(item, language_code),
                _item_size_for_language(item, language_code),
                qty_display,
                item.category,
            )
        )
    return rows


class RequirementHeaderListView(EventScopedListView):
    model = RequirementHeader
    template_name = "requirements/header_list.html"
    row_fields = ("form_number", "requirement_date", "volunteer_name", "get_route_area_display", "current_address", "chaturmas_place_address", "get_status_display")
    headers = ["Form No.", "Filled Date", "Janaar name", "Route", "હાલનું સરનામું", "ચાતુર્માસ સ્થળનું સરનામું", "Status"]
    search_fields = ["form_number", "order_number", "volunteer_name", "current_address", "chaturmas_place_address", "remarks"]
    create_url_name = "requirements:collect"
    edit_url_name = "requirements:collect-edit"
    delete_url_name = "requirements:header-delete"
    detail_url_name = "requirements:header-detail"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["page_title"] = "Requirements"
        context["page_subtitle"] = "Collect new orders, review saved orders, and edit them when needed."
        context["create_url"] = reverse_lazy(self.create_url_name)
        headers = list(self.object_list)
        context["summary"] = {
            "orders": len(headers),
            "items": sum(header.lines.count() for header in headers),
            "qty_total": sum((line.required_qty for header in headers for line in header.lines.all()), start=0),
        }
        return context

    def get_queryset(self):
        return super().get_queryset()

    def get_table_rows(self):
        rows = []
        for obj in self.object_list:
            row_kwargs = self.get_row_url_kwargs(obj)
            rows.append(
                {
                    "object": obj,
                    "cells": [
                        getattr(obj, field)() if callable(getattr(obj, field)) else getattr(obj, field, "")
                        for field in self.row_fields
                    ],
                    "view_url": reverse(self.detail_url_name, kwargs=row_kwargs) if self.detail_url_name else "",
                    "edit_url": reverse(self.edit_url_name, kwargs=row_kwargs) if self.edit_url_name else "",
                    "delete_url": reverse(self.delete_url_name, kwargs=row_kwargs) if self.delete_url_name else "",
                }
            )
        return rows


class RequirementHeaderExportView(LoginRequiredMixin, View):
    def get(self, request, *args, **kwargs):
        event_id = request.GET.get("event")
        event = Event.objects.filter(pk=event_id, is_active=True).first() if event_id else Event.objects.filter(is_current=True, is_active=True).first()
        if event is None:
            return HttpResponse("No active event found.", status=404)

        headers = (
            RequirementHeader.objects.filter(event=event)
            .select_related("upashray")
            .prefetch_related("lines__item")
            .exclude(status=RequirementStatus.DRAFT)
            .order_by("-updated_at", "-created_at")
        )

        workbook = Workbook()
        ws_all = workbook.active
        ws_all.title = "All Data"
        ws_analysis = workbook.create_sheet("Item Analysis")
        ws_summary = workbook.create_sheet("Order Summary")

        header_fill = PatternFill("solid", fgColor="DCE9F5")
        total_fill = PatternFill("solid", fgColor="FFF3BF")
        bold_font = Font(bold=True, color="14324F")
        center = Alignment(horizontal="center", vertical="center")
        right = Alignment(horizontal="right", vertical="center")

        all_headers = [
            "Form No.",
            "Order ID",
            "Requirement Date",
            "Upashray",
            "Status",
            "Locked",
            "Item Code",
            "Item Name",
            "Category",
            "Qty Required",
            "Remarks",
        ]
        ws_all.append(all_headers)
        for cell in ws_all[1]:
            cell.fill = header_fill
            cell.font = bold_font
            cell.alignment = center

        item_totals = {}
        order_totals = []
        for header in headers:
            order_qty = 0
            lines = list(header.lines.select_related("item").all())
            for line in lines:
                display_name = line.item.item_name_gu or line.item.item_name
                ws_all.append([
                    header.form_number or "",
                    header.order_number or "",
                    header.requirement_date.strftime("%d-%b-%Y") if header.requirement_date else "",
                    str(header.upashray),
                    header.get_status_display(),
                    "Yes" if header.is_locked else "No",
                    line.item.item_code,
                    display_name,
                    line.item.get_category_display(),
                    float(line.required_qty),
                    line.remarks,
                ])
                item_totals.setdefault(
                    line.item_id,
                    {
                        "Code": line.item.item_code,
                        "Item": display_name,
                        "Category": line.item.get_category_display(),
                        "Total Qty": 0,
                    },
                )
                item_totals[line.item_id]["Total Qty"] += float(line.required_qty)
                order_qty += float(line.required_qty)
            order_totals.append(
                {
                    "Form No.": header.form_number or "",
                    "Order ID": header.order_number or "",
                    "Requirement Date": header.requirement_date.strftime("%d-%b-%Y") if header.requirement_date else "",
                    "Upashray": str(header.upashray),
                    "Status": header.get_status_display(),
                    "Lines": len(lines),
                    "Total Qty": order_qty,
                }
            )

        if ws_all.max_row > 1:
            all_data_last_row = ws_all.max_row
            ws_all.append(["TOTAL", "", "", "", "", "", "", "", f"=SUM(I2:I{all_data_last_row})", ""])
            for cell in ws_all[ws_all.max_row]:
                cell.fill = total_fill
                cell.font = bold_font

        analysis_headers = ["Item Code", "Item Name", "Category", "Total Qty"]
        ws_analysis.append(analysis_headers)
        for cell in ws_analysis[1]:
            cell.fill = header_fill
            cell.font = bold_font
            cell.alignment = center
        for row in sorted(item_totals.values(), key=lambda data: data["Code"]):
            ws_analysis.append([row["Code"], row["Item"], row["Category"], row["Total Qty"]])
        analysis_last_row = ws_analysis.max_row
        ws_analysis.append(["TOTAL", "", "", f"=SUM(D2:D{analysis_last_row})"])
        for cell in ws_analysis[ws_analysis.max_row]:
            cell.fill = total_fill
            cell.font = bold_font

        summary_headers = ["Form No.", "Order ID", "Requirement Date", "Upashray", "Status", "Lines", "Total Qty"]
        ws_summary.append(summary_headers)
        for cell in ws_summary[1]:
            cell.fill = header_fill
            cell.font = bold_font
            cell.alignment = center
        for row in order_totals:
            ws_summary.append([row[h] for h in summary_headers])
        summary_last_row = ws_summary.max_row
        ws_summary.append(["TOTAL", "", "", "", "", f"=SUM(F2:F{summary_last_row})", f"=SUM(G2:G{summary_last_row})"])
        for cell in ws_summary[ws_summary.max_row]:
            cell.fill = total_fill
            cell.font = bold_font

        for ws in (ws_all, ws_analysis, ws_summary):
            ws.freeze_panes = "A2"
            for column_cells in ws.columns:
                max_len = 0
                column_letter = column_cells[0].column_letter
                for cell in column_cells:
                    try:
                        text = str(cell.value or "")
                    except Exception:
                        text = ""
                    max_len = max(max_len, len(text))
                ws.column_dimensions[column_letter].width = min(max_len + 2, 40)
            for row in ws.iter_rows(min_row=2):
                for cell in row:
                    if cell.column in (1, 2, 3, 4, 5):
                        cell.alignment = center
                    elif cell.column in (9, 10):
                        cell.alignment = right

        buffer = BytesIO()
        workbook.save(buffer)
        response = HttpResponse(
            buffer.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = 'attachment; filename="requirements.xlsx"'
        return response


class RequirementCollectionView(View):
    template_name = "requirements/collect.html"
    confirm_required_fields = (
        "route_area",
        "pujya_shri_name",
        "thana_count",
        "area",
        "current_address",
        "chaturmas_place_address",
        "requirement_date",
        "chaturmas_entry_date",
        "volunteer_name",
        "volunteer_mobile",
        "stay_type",
    )
    confirm_required_field_labels = {
        "pujya_shri_name": "પૂજ્ય શ્રી",
        "thana_count": "ઠાણા",
        "area": "વિસ્તાર",
        "current_address": "હાલનું સરનામું",
        "chaturmas_place_address": "ચાતુર્માસ સ્થળનું સરનામું",
        "requirement_date": "ફોર્મ તારીખ",
        "chaturmas_entry_date": "ચાતુર્માસ પ્રવેશ તારીખ",
        "volunteer_name": "જનારનું નામ",
        "volunteer_mobile": "જનારનું નંબર",
        "route_area": "રૂટ",
        "stay_type": "સંઘ ઉપાશ્રય / સ્થિરવાસ",
    }

    def _get_event(self):
        event_token = self.kwargs.get("event_token") or self.request.GET.get("event_token") or self.request.POST.get("event_token")
        if event_token:
            try:
                uuid.UUID(str(event_token))
            except (ValueError, AttributeError):
                return None
            return Event.objects.filter(public_form_token=event_token, is_active=True).first()
        event_id = self.request.GET.get("event") or self.request.POST.get("event") or self.kwargs.get("event_pk")
        if event_id:
            try:
                event_id = int(event_id)
            except (ValueError, TypeError):
                return None
            return Event.objects.filter(pk=event_id, is_active=True).first()
        return Event.objects.filter(is_current=True, is_active=True).first()

    def _is_public_flow(self):
        return bool(self.kwargs.get("event_token"))

    def _get_header(self, event):
        token = self.kwargs.get("token") or self.request.POST.get("token") or self.request.GET.get("token")
        if token:
            try:
                uuid.UUID(str(token))
            except (ValueError, AttributeError):
                return None
            return RequirementHeader.objects.filter(public_view_token=token, event=event).first()
        pk = self.kwargs.get("pk") or self.request.POST.get("header_pk") or self.request.GET.get("header_pk")
        if not pk or event is None:
            return None
        try:
            pk = int(pk)
        except (ValueError, TypeError):
            return None
        return RequirementHeader.objects.filter(pk=pk, event=event).first()

    def _editing_allowed(self, event, header, user):
        if header and header.status in (
            RequirementStatus.CONFIRMED,
            RequirementStatus.PACKED,
            RequirementStatus.DELIVERED,
            RequirementStatus.SUBMITTED,
            RequirementStatus.IN_PROGRESS,
            RequirementStatus.CLOSED,
            RequirementStatus.CANCELLED,
            RequirementStatus.RETURN_REQUESTED,
            RequirementStatus.RETURN_DONE,
            RequirementStatus.RECEIVED_BY_MS,
        ):
            return False
        return True

    def _get_items(self, event):
        base_items = list(
            Item.objects.filter(event=event, is_active=True, parent_item__isnull=True)
            .prefetch_related("variants")
            .order_by("standard_serial", "pk")
        )
        items = []
        for item in base_items:
            variants = list(item.variants.filter(is_active=True).order_by("variant_name", "pk"))
            if variants:
                items.extend(variants)
            else:
                items.append(item)
        return items

    def _resolve_upashray(self, event, upashray_name):
        name = (upashray_name or "").strip()
        if not name:
            existing = Upashray.objects.filter(event=event, is_active=True).order_by("name").first()
            if existing:
                return existing
            name = event.name
        upashray = Upashray.objects.filter(event=event, name__iexact=name, is_active=True).first()
        if upashray:
            return upashray
        return Upashray.objects.create(event=event, name=name)

    def _build_formset(self, items, data=None, initial_quantities=None):
        collection_formset = formset_factory(RequirementCollectionItemForm, extra=4)
        initial_quantities = initial_quantities or {}
        initial = []
        for item in items:
            initial.append(
                {
                    "item_id": item.pk,
                    "required_qty": initial_quantities.get(item.pk),
                }
            )
        return collection_formset(data=data, initial=initial, prefix="items")

    def _build_rows(self, items, formset, language_code, existing_quantities=None):
        rows = []
        existing_quantities = existing_quantities or {}
        variant_positions = {}
        for item in items:
            if item.parent_item_id:
                siblings = variant_positions.get(item.parent_item_id)
                if siblings is None:
                    siblings = list(item.parent_item.variants.filter(is_active=True).order_by("variant_name", "pk"))
                    variant_positions[item.parent_item_id] = siblings
        for item, form in zip(items, formset.forms, strict=False):
            serial = item.standard_serial or item.pk
            if item.parent_item_id:
                parent_serial = item.parent_item.standard_serial or item.parent_item.pk
                siblings = variant_positions.get(item.parent_item_id, [])
                suffix_index = next((idx for idx, sibling in enumerate(siblings) if sibling.pk == item.pk), 0)
                serial = f"{parent_serial}-{_variant_suffix(suffix_index)}"
            rows.append(
                {
                    "serial": serial,
                    "item": item,
                    "form": form,
                    "display_name": _item_name_for_language(item, language_code),
                    "item_size": _item_size_for_language(item, language_code),
                    "category_class": CATEGORY_ROW_CLASSES.get(item.category, "cat-general"),
                    "variant_row": bool(item.parent_item_id),
                    "ask_qty": bool(existing_quantities and item.pk not in existing_quantities),
                }
            )
        return rows

    def _group_rows(self, rows, language_code):
        grouped = []
        current_category = None
        current_group = None
        for row in rows:
            item = row["item"]
            if item.category != current_category:
                current_category = item.category
                current_group = {
                    "category": CATEGORY_LABELS.get(language_code, CATEGORY_LABELS["en"]).get(item.category, item.get_category_display()),
                    "category_class": CATEGORY_ROW_CLASSES.get(item.category, "cat-general"),
                    "rows": [],
                }
                grouped.append(current_group)
            current_group["rows"].append(row)
        return grouped

    def _build_context(self, request, event, header, form, formset, items, existing_quantities=None):
        language_code = _lang_code(request)
        draft_key = f"kmm.requirements.collect.{event.pk if event else 'noevent'}.{header.pk if header else 'new'}"
        is_public_flow = self._is_public_flow()
        note_values = []
        if header and header.remarks:
            note_values = header.remarks.splitlines()
        note_values = (note_values + ["", "", "", ""])[:4]
        return {
            "collect_base_template": "public/form_base.html" if is_public_flow else "base.html",
            "is_public_flow": is_public_flow,
            "event": event,
            "header": header,
            "form": form,
            "formset": formset,
            "extra_note_values": note_values,
            "item_groups": self._group_rows(self._build_rows(items, formset, language_code, existing_quantities), language_code),
            "language_code": language_code,
            "page_title": "જરૂરિયાતો એકત્ર કરો" if language_code == "gu" else "Collect Requirements",
            "page_subtitle": "જથ્થો ભરો. સાચવો અને પછી એક જ ઓર્ડર પર ફરીથી સંપાદિત કરો."
            if language_code == "gu"
            else "Fill quantities, save once, and edit the same order later.",
            "list_url": reverse_lazy("requirements:header-list"),
            "public_requests_url": reverse_lazy("public-requests"),
            "can_save": bool(event),
            "order_number": header.order_number if header else None,
            "public_collect_url": reverse("requirements:public-collect", kwargs={"event_token": event.public_form_token}) if event else None,
            "editing_allowed": self._editing_allowed(event, header, request.user),
            "event_requires_lock": bool(event and not event.allow_requirement_edit_after_confirm),
            "draft_storage_key": draft_key,
            "sub_route_data": json.dumps(RequirementHeader.SUB_ROUTE_CHOICES),
            "pending_edit_requests": header.edit_requests.filter(is_resolved=False).count() if header else 0,
        }

    def _render_summary_html(self, request, header):
        return render_to_string(
            "requirements/_saved_order_summary.html",
            {
                "request": request,
                "header": header,
                "lang": _lang_code(request),
            },
        )

    def get(self, request, *args, **kwargs):
        event = self._get_event()
        header = self._get_header(event) if event else None
        items = self._get_items(event) if event else []
        existing_quantities = {line.item_id: int(line.required_qty) for line in header.lines.all()} if header else {}
        form = RequirementCollectionHeaderForm(instance=header, current_event=event)
        formset = self._build_formset(items, initial_quantities=existing_quantities)
        return render(request, self.template_name, self._build_context(request, event, header, form, formset, items, existing_quantities))

    @transaction.atomic
    def post(self, request, *args, **kwargs):
        event = self._get_event()
        language_code = _lang_code(request)
        if event is None:
            messages.error(request, "સક્રિય ઇવેન્ટ મળી નથી." if language_code == "gu" else "No active event found.")
            return redirect(reverse_lazy("dashboard:home"))

        header = self._get_header(event)
        items = self._get_items(event)
        form = RequirementCollectionHeaderForm(request.POST, instance=header, current_event=event)
        existing_quantities = {line.item_id: int(line.required_qty) for line in header.lines.all()} if header else {}
        formset = self._build_formset(items, data=request.POST, initial_quantities=existing_quantities)

        save_details_now = "save_details" in request.POST
        confirm_now = "confirm" in request.POST
        editing_allowed = self._editing_allowed(event, header, request.user)

        if not editing_allowed:
            messages.error(request, "This confirmed order is locked by admin.")
            return render(request, self.template_name, self._build_context(request, event, header, form, formset, items))

        if header and header.is_locked and not request.user.is_superuser:
            messages.error(request, "This requirement order is locked.")
            return render(request, self.template_name, self._build_context(request, event, header, form, formset, items))

        if save_details_now:
            if not form.is_valid():
                if request.headers.get("x-requested-with") == "XMLHttpRequest":
                    payload = {
                        "ok": False,
                        "message": "Please fix the highlighted fields.",
                        "errors": form.errors.get_json_data(),
                    }
                    return JsonResponse(payload, status=400)
                return render(request, self.template_name, self._build_context(request, event, header, form, formset, items, existing_quantities))

            header_obj = form.save(commit=False)
            header_obj.event = event
            header_obj.requirement_date = form.cleaned_data.get("requirement_date") or timezone.localdate()
            header_obj.upashray = self._resolve_upashray(event, form.cleaned_data.get("upashray_name"))
            if header_obj.upashray is None:
                form.add_error("upashray_name", "Upashray name is required.")
                return render(request, self.template_name, self._build_context(request, event, header, form, formset, items, existing_quantities))

            old_status = header.status if header else None
            header_obj.status = header.status if header else RequirementStatus.NOT_CONFIRMED
            if old_status == RequirementStatus.DRAFT:
                header_obj.status = RequirementStatus.NOT_CONFIRMED
            header_obj.order_number = header.order_number if header else None
            header_obj.is_locked = header.is_locked if header else False
            header_obj.locked_at = header.locked_at if header else None
            header_obj.remarks = "\n".join(
                [
                    (request.POST.get("extra_note_1") or "").strip(),
                    (request.POST.get("extra_note_2") or "").strip(),
                    (request.POST.get("extra_note_3") or "").strip(),
                    (request.POST.get("extra_note_4") or "").strip(),
                ]
            ).strip()
            header_obj.save()

            item_errors = []
            if formset.is_valid():
                RequirementLine.objects.filter(event=event, requirement=header_obj).delete()
                seen_item_ids = set()
                for item_form in formset.cleaned_data:
                    item_id = item_form.get("item_id")
                    qty = item_form.get("required_qty") or 0
                    if not item_id or qty <= 0:
                        continue
                    if item_id in seen_item_ids:
                        continue
                    seen_item_ids.add(item_id)
                    RequirementLine.objects.create(
                        event=event,
                        requirement=header_obj,
                        item_id=item_id,
                        required_qty=qty,
                    )
            else:
                item_errors = formset.errors

            if request.headers.get("x-requested-with") == "XMLHttpRequest":
                return JsonResponse(
                    {
                        "ok": True,
                        "message": "Saved",
                        "header_pk": header_obj.pk,
                        "detail": "Basic details saved. You can continue filling items.",
                        "summary_html": self._render_summary_html(request, header_obj),
                        "item_errors": item_errors,
                    },
                )
            messages.success(request, "Basic details saved. You can continue filling items.")
            saved_existing_quantities = {line.item_id: int(line.required_qty) for line in header_obj.lines.all()}
            return render(
                request,
                self.template_name,
                self._build_context(request, event, header_obj, form, formset, items, saved_existing_quantities),
            )

        if not form.is_valid() or not formset.is_valid():
            if request.headers.get("x-requested-with") == "XMLHttpRequest":
                payload = {
                    "ok": False,
                    "message": "Please fix the highlighted fields.",
                    "errors": form.errors.get_json_data(),
                }
                if not formset.is_valid():
                    payload["item_errors"] = formset.errors
                return JsonResponse(payload, status=400)
            return render(request, self.template_name, self._build_context(request, event, header, form, formset, items, existing_quantities))

        header_obj = form.save(commit=False)
        header_obj.event = event
        header_obj.requirement_date = form.cleaned_data.get("requirement_date") or timezone.localdate()
        header_obj.upashray = self._resolve_upashray(event, form.cleaned_data.get("upashray_name"))
        if header_obj.upashray is None:
            form.add_error("upashray_name", "Upashray name is required.")
            return render(request, self.template_name, self._build_context(request, event, header, form, formset, items, existing_quantities))
        if confirm_now:
            missing_fields = []
            for field_name in self.confirm_required_fields:
                value = form.cleaned_data.get(field_name)
                if value in (None, "", []):
                    form.add_error(field_name, "This field is required.")
                    missing_fields.append(field_name)
            if missing_fields:
                if request.headers.get("x-requested-with") == "XMLHttpRequest":
                    return JsonResponse({"ok": False, "message": "Missing required fields: " + ", ".join(missing_fields)}, status=400)
                missing_labels = [self.confirm_required_field_labels.get(name, name) for name in missing_fields]
                messages.error(
                    request,
                    "આ વિગતો જરૂરી છે: " + ", ".join(missing_labels) if language_code == "gu" else "Missing required details: " + ", ".join(missing_labels),
                )
                return render(
                    request,
                    self.template_name,
                    self._build_context(request, event, header, form, formset, items, existing_quantities),
                )
            header_obj.status = RequirementStatus.CONFIRMED
        else:
            old_status = header.status if header else None
            header_obj.status = header.status if header else RequirementStatus.NOT_CONFIRMED
            if old_status == RequirementStatus.DRAFT:
                header_obj.status = RequirementStatus.NOT_CONFIRMED
            header_obj.order_number = header.order_number if header else None
            header_obj.is_locked = header.is_locked if header else False
            header_obj.locked_at = header.locked_at if header else None
        header_obj.remarks = "\n".join(
            [
                (request.POST.get("extra_note_1") or "").strip(),
                (request.POST.get("extra_note_2") or "").strip(),
                (request.POST.get("extra_note_3") or "").strip(),
                (request.POST.get("extra_note_4") or "").strip(),
            ]
        ).strip()
        header_obj.save()

        RequirementLine.objects.filter(event=event, requirement=header_obj).delete()
        seen_item_ids = set()
        for item_form in formset.cleaned_data:
            item_id = item_form.get("item_id")
            qty = item_form.get("required_qty") or 0
            if not item_id or qty <= 0:
                continue
            if item_id in seen_item_ids:
                continue
            seen_item_ids.add(item_id)
            RequirementLine.objects.create(
                event=event,
                requirement=header_obj,
                item_id=item_id,
                required_qty=qty,
            )

        if confirm_now:
            send_form_backup_email(header_obj)
            form_label = header_obj.form_number or header_obj.order_number or "-"
            messages.success(request, f"Requirement sent to team. Form No: {form_label}")
            if self._is_public_flow():
                return redirect(reverse("public-requests"))
            return redirect(reverse("requirements:header-list"))
        messages.success(request, "Data saved. Press Confirm to send Requirement to team.")
        return render(request, self.template_name, self._build_context(request, event, header_obj, form, formset, items, existing_quantities))


class RequirementCollectionPrintView(View):
    def get(self, request, pk=None, token=None):
        query = RequirementHeader.objects.select_related("event", "upashray")
        if token is not None:
            header = query.filter(public_view_token=token).first()
        else:
            header = query.filter(pk=pk).first()
        if header is None:
            messages.error(request, "Requirement order not found.")
            return redirect(reverse_lazy("requirements:header-list"))

        items = list(
            RequirementLine.objects.filter(requirement=header)
            .select_related("item", "item__parent_item")
        )
        language_code = _lang_code(request)
        if language_code == "gu":
            gujarati_mode = request.GET.get("type", "opted")
            if gujarati_mode == "full":
                all_items = []
                base_items = Item.objects.filter(
                    event=header.event, is_active=True, parent_item__isnull=True
                ).prefetch_related("variants").order_by("standard_serial", "pk")
                for item in base_items:
                    variants = list(item.variants.filter(is_active=True).order_by("variant_name", "pk"))
                    if variants:
                        all_items.extend(variants)
                    else:
                        all_items.append(item)
                qty_map = {line.item_id: line.required_qty for line in items}
                line_rows = []
                for item in all_items:
                    qty = qty_map.get(item.pk, 0)
                    qty_display = "--" if not qty or qty <= 0 else _format_qty(qty)
                    line_rows.append((
                        _line_serial_display(item),
                        _item_name_for_language(item, "gu"),
                        _item_size_for_language(item, "gu"),
                        qty_display,
                        item.category,
                    ))
                contact = _format_main_contact(header.event)
                contact_gu = contact.replace("Main Event Manager:", "મુખ્ય સંપર્ક:")
                return generate_gujarati_pdf_fpdf2(
                    header, line_rows, contact_gu,
                    filename=f'{header.order_number or "requirement-order"}-full.pdf',
                )
            line_rows = _requirement_pdf_rows(items, "gu")
            contact = _format_main_contact(header.event)
            contact_gu = contact.replace("Main Event Manager:", "મુખ્ય સંપર્ક:")
            return generate_gujarati_pdf_fpdf2(
                header, line_rows, contact_gu,
                filename=f'{header.order_number or "requirement-order"}-gujarati.pdf',
            )
        return self._render_pdf(header, items, "en")

    def _render_pdf_gujarati_html(self, header, lines):
        line_rows = _requirement_pdf_rows(lines, "gu")
        half = (len(line_rows) + 1) // 2
        left_lines = line_rows[:half]
        right_lines = line_rows[half:]
        max_rows = max(len(left_lines), len(right_lines))
        item_rows = []
        for idx in range(max_rows):
            left = left_lines[idx] if idx < len(left_lines) else ("", "", "", "")
            right = right_lines[idx] if idx < len(right_lines) else ("", "", "", "")
            item_rows.append((left, right))

        resolved_font_path = _resolve_font_path()
        font_src = resolved_font_path.replace("\\", "/")
        pdf_font_url = f"file:///{font_src.lstrip('/')}"

        context = {
            "header": header,
            "rows": item_rows,
            "extra_note_values": ((header.remarks.splitlines() if header.remarks else []) + ["", "", "", ""])[:4],
            "contact": _format_main_contact(header.event),
            "order_number": header.order_number,
            "lang": "gu",
            "logo_exists": (Path(settings.BASE_DIR) / "pdf_header.png").exists(),
            "pdf_header_path": str((Path(settings.BASE_DIR) / "pdf_header.png").resolve()).replace("\\", "/"),
            "pdf_font_url": pdf_font_url,
            "pdf_font_path": resolved_font_path,
        }

        return generate_weasyprint_pdf(
            "requirements/gujarati_pdf.html",
            context,
            filename=f'{header.order_number or "requirement-order"}-gujarati.pdf',
            extra_css="""
                @page { size: A4; margin: 8mm 8mm 10mm 8mm; }
            """,
            font_path=resolved_font_path,
        )

    def _render_pdf_gujarati(self, header, lines):
        buffer = BytesIO()
        pdf_font_name = PDF_GUJARATI_FONT_NAME
        document = SimpleDocTemplate(
            buffer,
            pagesize=A4,
            leftMargin=8 * mm,
            rightMargin=8 * mm,
            topMargin=8 * mm,
            bottomMargin=10 * mm,
        )

        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            "KMMTitle",
            parent=styles["Title"],
            fontName=pdf_font_name,
            fontSize=15,
            leading=17,
            textColor=colors.HexColor("#14324f"),
            alignment=TA_LEFT,
            spaceAfter=2,
        )
        heading_style = ParagraphStyle(
            "KMMHeading",
            parent=styles["Heading2"],
            fontName=pdf_font_name,
            fontSize=10,
            leading=12,
            textColor=colors.HexColor("#14324f"),
            alignment=TA_LEFT,
        )
        body_style = ParagraphStyle(
            "KMMBody",
            parent=styles["BodyText"],
            fontName=pdf_font_name,
            fontSize=8,
            leading=9.2,
            alignment=TA_LEFT,
        )
        small_style = ParagraphStyle(
            "KMMSmall",
            parent=styles["BodyText"],
            fontName=pdf_font_name,
            fontSize=7.2,
            leading=8.0,
            alignment=TA_LEFT,
        )

        def p(text, style=body_style):
            parts = [escape(part) for part in str(text or "-").splitlines() or ["-"]]
            return Paragraph("<br/>".join(parts), style)

        base_url = getattr(settings, "PUBLIC_SITE_URL", "").rstrip("/") or "http://127.0.0.1:8000"
        qr_value = f"{base_url}{reverse('requirements:header-detail', kwargs={'pk': header.pk})}"
        qr_drawing = createBarcodeDrawing("QR", value=qr_value, barBorder=0, width=22 * mm, height=22 * mm)

        logo_path = Path(settings.BASE_DIR) / "pdf_header.png"
        logo_image = Image(str(logo_path), width=63 * mm, height=19.6 * mm) if logo_path.exists() else None
        brand_content = logo_image if logo_image else Paragraph("કલ્યાણ મિત્ર મંડળ", title_style)
        brand_row = Table([[brand_content, qr_drawing]], colWidths=[156 * mm, 22 * mm])
        brand_row.setStyle(
            TableStyle(
                [
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                    ("ALIGN", (1, 0), (1, 0), "RIGHT"),
                    ("LEFTPADDING", (0, 0), (-1, -1), 0),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 0),
                    ("TOPPADDING", (0, 0), (-1, -1), 0),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
                ]
            )
        )

        story = [
            brand_row,
            Spacer(1, 2),
            Paragraph(f"વૈયાવચ્ચ લાભ નંબર: {header.form_number or header.order_number}", heading_style),
            Spacer(1, 4),
        ]

        status_table = Table(
            [[
                p("[ ] બધી વસ્તુઓ પેક થઈ", small_style),
                p("[ ] વિતરણ માટે તૈયાર", small_style),
                p("ચેક કરનાર: ____________________", small_style),
            ]],
            colWidths=[54 * mm, 54 * mm, 76 * mm],
        )
        status_table.setStyle(
            TableStyle(
                [
                    ("BOX", (0, 0), (-1, -1), 0.6, colors.HexColor("#14324f")),
                    ("INNERGRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#9bb4c9")),
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                    ("FONTNAME", (0, 0), (-1, -1), pdf_font_name),
                    ("LEADING", (0, 0), (-1, -1), 8.5),
                ]
            )
        )

        basic_rows = [
            [p("પૂજ્ય શ્રી", small_style), p(header.pujya_shri_name, body_style), p("ઠાણા", small_style), p(header.thana_count, body_style)],
            [p("વિસ્તાર", small_style), p(header.area, body_style), p("ફોર્મ તારીખ", small_style), p(_format_pdf_date(header.requirement_date), body_style)],
            [p("હાલનું સરનામું", small_style), p(header.current_address, body_style), p("ચાતુર્માસ સ્થળનું સરનામું", small_style), p(header.chaturmas_place_address, body_style)],
            [p("ચાતુર્માસ પ્રવેશ તારીખ", small_style), p(_format_pdf_date(header.chaturmas_entry_date), body_style), p("જનારનું નામ", small_style), p(header.volunteer_name, body_style)],
            [p("સંઘ ઉપાશ્રય / સ્થિરવાસ", small_style), p(header.get_stay_type_display(), body_style), p("પૂજ્ય શ્રી સંભાળ લેનારનું નામ અને મોબાઇલ", small_style), p(f"{header.caretaker_name or '-'} {header.caretaker_mobile or ''}".strip(), body_style)],
        ]
        basic_table = Table(basic_rows, colWidths=[40 * mm, 53 * mm, 42 * mm, 49 * mm])
        story.extend([status_table, Spacer(1, 6), basic_table, Spacer(1, 6)])

        paired_rows = []
        line_rows = _requirement_pdf_rows(lines, "gu")
        half = (len(line_rows) + 1) // 2
        left_lines = line_rows[:half]
        right_lines = line_rows[half:]
        max_rows = max(len(left_lines), len(right_lines))

        paired_rows.append([
            p("નં.", small_style), p("વસ્તુનું નામ", small_style), p("પ્રકાર/સાઈઝ", small_style), p("નંગ", small_style),
            p("નં.", small_style), p("વસ્તુનું નામ", small_style), p("પ્રકાર/સાઈઝ", small_style), p("નંગ", small_style),
        ])

        def row_bits(row):
            if not row:
                return [p("", small_style), p("", small_style), p("", small_style), p("", small_style)]
            return [
                p(row[0], small_style),
                p(row[1], body_style),
                p(row[2], small_style),
                p(row[3], body_style),
            ]

        for idx in range(max_rows):
            left = left_lines[idx] if idx < len(left_lines) else None
            right = right_lines[idx] if idx < len(right_lines) else None
            paired_rows.append(row_bits(left) + row_bits(right))

        item_table = Table(paired_rows, colWidths=[10 * mm, 45 * mm, 30 * mm, 12 * mm, 10 * mm, 45 * mm, 30 * mm, 12 * mm], repeatRows=1)
        item_table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#dce9f5")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.HexColor("#14324f")),
                    ("BOX", (0, 0), (-1, -1), 0.6, colors.HexColor("#14324f")),
                    ("INNERGRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#9bb4c9")),
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                    ("FONTNAME", (0, 0), (-1, -1), pdf_font_name),
                    ("LEADING", (0, 0), (-1, -1), 8.5),
                ]
            )
        )
        story.append(item_table)
        note_rows = ((header.remarks.splitlines() if header.remarks else []) + ["", "", "", ""])[:4]
        story.append(Spacer(1, 4))
        story.append(Paragraph("Extra Items", heading_style))
        note_table = Table([[p(note, body_style)] for note in note_rows], colWidths=[180 * mm])
        note_table.setStyle(
            TableStyle(
                [
                    ("BOX", (0, 0), (-1, -1), 0.6, colors.HexColor("#14324f")),
                    ("INNERGRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#9bb4c9")),
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                    ("FONTNAME", (0, 0), (-1, -1), pdf_font_name),
                ]
            )
        )
        story.append(note_table)

        footer_contact = _format_main_contact(header.event)

        class NumberedCanvas(SharedNumberedCanvas):
            def draw_footer(self, page_count):
                self.saveState()
                self.setStrokeColor(colors.HexColor("#9bb4c9"))
                self.setLineWidth(0.5)
                self.line(8 * mm, 11 * mm, A4[0] - 8 * mm, 11 * mm)
                self.setFont(pdf_font_name, 7.5)
                self.setFillColor(colors.HexColor("#14324f"))
                self.drawString(8 * mm, 5.2 * mm, footer_contact)
                self.drawRightString(A4[0] - 8 * mm, 5.2 * mm, f"Page {self._pageNumber} of {page_count}")
                self.restoreState()

        document.build(story, canvasmaker=NumberedCanvas)
        pdf_data = buffer.getvalue()
        buffer.close()

        response = HttpResponse(content_type="application/pdf")
        response["Content-Disposition"] = f'attachment; filename="{header.order_number or "requirement-order"}-gujarati.pdf"'
        response.write(pdf_data)
        return response

    def _render_pdf(self, header, lines, language_code):
        buffer = BytesIO()
        pdf_font_name = PDF_GUJARATI_FONT_NAME if PDF_GUJARATI_FONT_NAME != "Helvetica" else PDF_FONT_NAME
        document = SimpleDocTemplate(
            buffer,
            pagesize=A4,
            leftMargin=12 * mm,
            rightMargin=12 * mm,
            topMargin=12 * mm,
            bottomMargin=12 * mm,
        )

        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            "KMMTitle",
            parent=styles["Title"],
            fontName=pdf_font_name,
            fontSize=18,
            leading=22,
            textColor=colors.HexColor("#14324f"),
            spaceAfter=6,
            alignment=TA_LEFT,
        )
        heading_style = ParagraphStyle(
            "KMMHeading",
            parent=styles["Heading2"],
            fontName=pdf_font_name,
            fontSize=11,
            leading=14,
            textColor=colors.HexColor("#14324f"),
            alignment=TA_LEFT,
        )
        body_style = ParagraphStyle(
            "KMMBody",
            parent=styles["BodyText"],
            fontName=pdf_font_name,
            fontSize=9,
            leading=12,
            alignment=TA_LEFT,
        )
        small_style = ParagraphStyle(
            "KMMSmall",
            parent=styles["BodyText"],
            fontName=pdf_font_name,
            fontSize=8,
            leading=10,
            alignment=TA_LEFT,
        )

        def p(text, style=body_style):
            parts = [escape(part) for part in str(text or "-").splitlines() or ["-"]]
            return Paragraph("<br/>".join(parts), style)

        story = [
            Paragraph("Requirement Order" if language_code != "gu" else "àªµà«ˆàª¯àª¾àªµàªšà«àªš àª²àª¾àª­ àªªàª¤à«àª°àª•", title_style),
            Paragraph(f"Order No.: {header.order_number}", heading_style),
            Spacer(1, 4),
        ]

        base_url = getattr(settings, "PUBLIC_SITE_URL", "").rstrip("/")
        if not base_url:
            base_url = "http://127.0.0.1:8000"
        qr_value = f"{base_url}{reverse('requirements:header-detail', kwargs={'pk': header.pk})}"
        qr_drawing = createBarcodeDrawing("QR", value=qr_value, barBorder=0, width=26 * mm, height=26 * mm)
        header_strip = Table(
            [
                [
                    [
                        Paragraph("Requirement Order" if language_code != "gu" else "Ã ÂªÅ“Ã ÂªÂ°Ã Â«â€šÃ ÂªÂ°Ã ÂªÂ¿Ã ÂªÂ¯Ã ÂªÂ¾Ã ÂªÂ¤ Ã Âªâ€œÃ ÂªÂ°Ã Â«ÂÃ ÂªÂ¡Ã ÂªÂ°", title_style),
            Paragraph(f"Form No.: {header.form_number or header.order_number}", heading_style),
                    ],
                    qr_drawing,
                ]
            ],
            colWidths=[150 * mm, 26 * mm],
        )
        header_strip.setStyle(
            TableStyle(
                [
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                    ("ALIGN", (1, 0), (1, 0), "RIGHT"),
                    ("LEFTPADDING", (0, 0), (-1, -1), 0),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 0),
                ]
            )
        )
        logo_path = Path(settings.BASE_DIR) / "pdf_header.png"
        logo_image = Image(str(logo_path), width=90 * mm, height=26.5 * mm) if logo_path.exists() else None
        gujarati_title_style = ParagraphStyle(
            "KMMGujaratiTitle",
            parent=title_style,
            fontName=PDF_GUJARATI_FONT_NAME,
        )
        brand_title = Paragraph("àª•àª²à«àª¯àª¾àª£ àª®àª¿àª¤à«àª° àª®àª‚àª¡àª³", gujarati_title_style)
        display_no = header.form_number or header.order_number or "-"
        order_label_text = f"Vaiyavachch laabh number: {display_no}" if language_code != "gu" else f"àªµà«ˆàª¯àª¾àªµàªšà«àªš àª²àª¾àª­ àª¨àª‚àª¬àª°: {display_no}"
        order_label = Paragraph(order_label_text, heading_style)
        if logo_image:
            brand_row = Table([[logo_image, qr_drawing]], colWidths=[150 * mm, 26 * mm])
        else:
            brand_row = Table([[brand_title, qr_drawing]], colWidths=[150 * mm, 26 * mm])
        brand_row.setStyle(
            TableStyle(
                [
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                    ("ALIGN", (1, 0), (1, 0), "RIGHT"),
                    ("LEFTPADDING", (0, 0), (-1, -1), 0),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 0),
                    ("TOPPADDING", (0, 0), (-1, -1), 0),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
                ]
            )
        )
        story = [brand_row, order_label, Spacer(1, 4)]

        status_table = Table(
            [
                [
                    p("[ ] All Items Packed" if language_code != "gu" else "[ ] àª¬àª§à«€ àªµàª¸à«àª¤à«àª“ àªªà«‡àª• àª¥àªˆ", small_style),
                    p("[ ] Ready for Distribution" if language_code != "gu" else "[ ] àªµàª¿àª¤àª°àª£ àª®àª¾àªŸà«‡ àª¤à«ˆàª¯àª¾àª°", small_style),
                    p("Checked by: ____________________" if language_code != "gu" else "àªšà«‡àª• àª•àª°àª¨àª¾àª°: ____________________", small_style),
                ]
            ],
            colWidths=[55 * mm, 55 * mm, 74 * mm],
        )
        status_table.setStyle(
            TableStyle(
                [
                    ("BOX", (0, 0), (-1, -1), 0.6, colors.HexColor("#14324f")),
                    ("INNERGRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#9bb4c9")),
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                    ("FONTNAME", (0, 0), (-1, -1), pdf_font_name),
                    ("LEADING", (0, 0), (-1, -1), 10),
                ]
            )
        )

        basic_rows = [
            [
                p("Pujya Shri" if language_code != "gu" else "àªªà«‚àªœà«àª¯ àª¶à«àª°à«€", small_style),
                p(header.pujya_shri_name, body_style),
                p("Current Address" if language_code != "gu" else "àªµàª°à«àª¤àª®àª¾àª¨ àª¸àª°àª¨àª¾àª®à«àª‚", small_style),
                p(header.current_address, body_style),
            ],
            [
                p("Thana" if language_code != "gu" else "àª¥àª¾àª£àª¾", small_style),
                p(header.thana_count, body_style),
                p("Area" if language_code != "gu" else "àªµàª¿àª¸à«àª¤àª¾àª°", small_style),
                p(header.area, body_style),
            ],
            [
                p("Chaturmas Place Address" if language_code != "gu" else "àªšàª¾àª¤à«àª°à«àª®àª¾àª¸ àª¸à«àª¥àª³ àª¸àª°àª¨àª¾àª®à«àª‚", small_style),
                p(header.chaturmas_place_address, body_style),
                p("Chaturmas Entry Date" if language_code != "gu" else "àªšàª¾àª¤à«àª°à«àª®àª¾àª¸ àªªà«àª°àªµà«‡àª¶ àª¤àª¾àª°à«€àª–", small_style),
                p(_format_pdf_date(header.chaturmas_entry_date), body_style),
            ],
            [
                p("Volunteer Name" if language_code != "gu" else "àªµà«‹àª²àª¨à«àªŸàª¿àª¯àª° àª¨àª¾àª®", small_style),
                p(header.volunteer_name, body_style),
                p("Stay Type" if language_code != "gu" else "àª°àª¹à«‡àª àª¾àª£ àªªà«àª°àª•àª¾àª°", small_style),
                p(header.get_stay_type_display(), body_style),
            ],
            [
                p("Care Taker Name" if language_code != "gu" else "àª¸àª‚àª­àª¾àª³àª¨àª¾àª° àª¨àª¾àª®", small_style),
                p(header.caretaker_name, body_style),
                p("Care Taker Contact" if language_code != "gu" else "àª¸àª‚àª­àª¾àª³àª¨àª¾àª° àª¸àª‚àªªàª°à«àª•", small_style),
                p(header.caretaker_mobile, body_style),
            ],
            [
                p("Special Request / Note" if language_code != "gu" else "àª–àª¾àª¸ àªµàª¿àª¨àª‚àª¤à«€ / àª¨à«‹àª‚àª§", small_style),
                p(header.remarks, body_style),
                p("", small_style),
                p("", body_style),
            ],
        ]
        basic_table = Table(basic_rows, colWidths=[28 * mm, 62 * mm, 35 * mm, 53 * mm])
        basic_table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, -1), colors.whitesmoke),
                    ("BOX", (0, 0), (-1, -1), 0.6, colors.HexColor("#14324f")),
                    ("INNERGRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#9bb4c9")),
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                    ("LEADING", (0, 0), (-1, -1), 12),
                    ("FONTNAME", (0, 0), (-1, -1), pdf_font_name),
                ]
            )
        )
        story.append(status_table)
        story.append(Spacer(1, 8))
        story.extend([basic_table, Spacer(1, 8), Paragraph("Item List" if language_code != "gu" else "àª†àª‡àªŸàª® àª¸à«‚àªšàª¿", heading_style)])

        table_data = [
            [
                p("Pack" if language_code != "gu" else "àªªà«‡àª•", small_style),
                p("Sr. No." if language_code != "gu" else "àª•à«àª°àª® àª¨àª‚.", small_style),
                p("Name of Item" if language_code != "gu" else "àªµàª¸à«àª¤à«àª¨à«àª‚ àª¨àª¾àª®", small_style),
                p("QTY Req" if language_code != "gu" else "àªœàª¥à«àª¥à«‹", small_style),
                p("Remark" if language_code != "gu" else "àª¨à«‹àª‚àª§", small_style),
            ]
        ]
        for line in lines:
            item = line.item
            display_name = _item_name_for_language(item, language_code)
            content = display_name
            if item.default_size:
                content = f"{display_name} ({item.default_size})"
            table_data.append([p("[ ]", small_style), p(item.standard_serial or item.pk), p(content), p(_format_qty(line.required_qty), body_style), p("")])

        item_table = Table(table_data, colWidths=[14 * mm, 18 * mm, 78 * mm, 20 * mm, 48 * mm], repeatRows=1)
        item_table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#dce9f5")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.HexColor("#14324f")),
                    ("BOX", (0, 0), (-1, -1), 0.6, colors.HexColor("#14324f")),
                    ("INNERGRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#9bb4c9")),
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                    ("FONTNAME", (0, 0), (-1, -1), pdf_font_name),
                    ("LEADING", (0, 0), (-1, -1), 12),
                ]
            )
        )
        story.append(item_table)
        note_rows = ((header.remarks.splitlines() if header.remarks else []) + ["", "", "", ""])[:4]
        story.append(Spacer(1, 4))
        story.append(Paragraph("Extra Items" if language_code != "gu" else "àªàª•à«àª¸à«àªŸà«àª°àª¾ àªµàª¸à«àª¤à«àª“", heading_style))
        note_table = Table([[p(note, body_style)] for note in note_rows], colWidths=[160 * mm])
        note_table.setStyle(
            TableStyle(
                [
                    ("BOX", (0, 0), (-1, -1), 0.6, colors.HexColor("#14324f")),
                    ("INNERGRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#9bb4c9")),
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                    ("FONTNAME", (0, 0), (-1, -1), pdf_font_name),
                ]
            )
        )
        story.append(note_table)

        footer_contact = _format_main_contact(header.event)

        class NumberedCanvas(SharedNumberedCanvas):
            def draw_footer(self, page_count):
                self.saveState()
                self.setStrokeColor(colors.HexColor("#9bb4c9"))
                self.setLineWidth(0.5)
                self.line(12 * mm, 14 * mm, A4[0] - 12 * mm, 14 * mm)
                self.setFont(pdf_font_name, 8)
                self.setFillColor(colors.HexColor("#14324f"))
                self.drawString(12 * mm, 8 * mm, footer_contact)
                self.drawRightString(A4[0] - 12 * mm, 8 * mm, f"Page {self._pageNumber} of {page_count}")
                self.restoreState()

        document.build(story, canvasmaker=NumberedCanvas)
        pdf_data = buffer.getvalue()
        buffer.close()

        response = HttpResponse(content_type="application/pdf")
        suffix = "-gujarati" if language_code == "gu" else ""
        response["Content-Disposition"] = f'attachment; filename="{header.order_number or "requirement-order"}{suffix}.pdf"'
        response.write(pdf_data)
        return response


class RequirementCollectionDetailView(View):
    def get(self, request, pk=None, token=None):
        query = RequirementHeader.objects.select_related("event", "upashray")
        if token is not None:
            header = query.filter(public_view_token=token).first()
        else:
            header = query.filter(pk=pk).first()
        if header is None:
            return HttpResponse("Requirement order not found.", status=404)
        lines = list(
            RequirementLine.objects.filter(requirement=header)
            .select_related("item")
            .order_by("item__parent_item__standard_serial", "item__standard_serial", "item__pk")
        )
        is_admin = request.user.is_superuser or request.user.groups.filter(name="KMM Admin").exists() or is_manager(request.user)
        edit_requests = list(header.edit_requests.filter(is_resolved=False).order_by("-created_at"))
        view_control = ViewControl.objects.filter(event=header.event).first()
        if request.user.is_authenticated:
            view_control = None
        return render(
            request,
            "requirements/header_detail.html" if request.headers.get("x-requested-with") == "XMLHttpRequest" else "requirements/header_detail_page.html",
            {
                "header": header,
                "lines": lines,
                "grouped_lines": _group_requirement_lines(lines),
                "lang": _lang_code(request),
                "is_admin": is_admin,
                "edit_requests": edit_requests,
                "view_control": view_control,
            },
        )


class ViewControlView(LoginRequiredMixin, View):
    def get_event(self, request):
        event_id = request.GET.get("event") or request.POST.get("event")
        if event_id:
            try:
                return Event.objects.filter(pk=int(event_id), is_active=True).first()
            except (ValueError, TypeError):
                pass
        return Event.objects.filter(is_current=True, is_active=True).first()

    def get(self, request):
        event = self.get_event(request)
        if not event:
            messages.error(request, "No active event found.")
            return redirect("dashboard:home")
        vc, _ = ViewControl.objects.get_or_create(event=event)
        form = ViewControlForm(instance=vc)
        return render(request, "requirements/view_control.html", {
            "form": form,
            "event": event,
            "lang": _lang_code(request),
        })

    def post(self, request):
        event = self.get_event(request)
        if not event:
            messages.error(request, "No active event found.")
            return redirect("dashboard:home")
        vc, _ = ViewControl.objects.get_or_create(event=event)
        form = ViewControlForm(request.POST, instance=vc)
        if form.is_valid():
            form.save()
            messages.success(request, "View control settings saved.")
            return redirect("requirements:view-control")
        return render(request, "requirements/view_control.html", {
            "form": form,
            "event": event,
            "lang": _lang_code(request),
        })


class RequirementStatusTransitionView(LoginRequiredMixin, View):
    ALLOWED_TRANSITIONS = {
        RequirementStatus.CONFIRMED: {
            "next_status": RequirementStatus.PACKED,
            "button_label": "Packing Done",
        },
        RequirementStatus.PACKED: {
            "next_status": RequirementStatus.DELIVERED,
            "button_label": "Delivery Done",
        },
        RequirementStatus.SUBMITTED: {
            "next_status": RequirementStatus.IN_PROGRESS,
            "button_label": "Packing Done",
        },
        RequirementStatus.IN_PROGRESS: {
            "next_status": RequirementStatus.CLOSED,
            "button_label": "Delivery Done",
        },
        RequirementStatus.CLOSED: {
            "next_status": RequirementStatus.RECEIVED_BY_MS,
            "button_label": "Received by M.S.",
        },
    }

    def post(self, request, pk):
        header = get_object_or_404(RequirementHeader, pk=pk)
        transition = self.ALLOWED_TRANSITIONS.get(header.status)
        if not transition:
            messages.error(request, "Status transition not allowed from current state.")
            return redirect("requirements:header-detail", pk=pk)
        header.status = transition["next_status"]
        if header.status in (RequirementStatus.IN_PROGRESS, RequirementStatus.PACKED):
            header.is_locked = True
            header.locked_at = timezone.now()
        header.updated_at = timezone.now()
        header.save(update_fields=["status", "is_locked", "locked_at", "updated_at"])
        messages.success(request, f"Status updated to {header.get_status_display()}.")
        return redirect("requirements:header-detail", pk=pk)


class RequirementUnlockView(LoginRequiredMixin, View):
    def post(self, request, pk):
        if not request.user.is_superuser and not request.user.groups.filter(name="KMM Admin").exists():
            messages.error(request, "Only admin/manager can unlock/lock forms.")
            return redirect("requirements:header-detail", pk=pk)
        header = get_object_or_404(RequirementHeader, pk=pk)
        header.is_locked = False
        header.locked_at = None
        header.updated_at = timezone.now()
        header.save(update_fields=["is_locked", "locked_at", "updated_at"])
        messages.success(request, "Form unlocked. Edit and lock again after updating.")
        return redirect("requirements:header-detail", pk=pk)


class RequirementLockView(LoginRequiredMixin, View):
    def post(self, request, pk):
        if not request.user.is_superuser and not request.user.groups.filter(name="KMM Admin").exists():
            messages.error(request, "Only admin/manager can unlock/lock forms.")
            return redirect("requirements:header-detail", pk=pk)
        header = get_object_or_404(RequirementHeader, pk=pk)
        header.is_locked = True
        header.locked_at = timezone.now()
        header.updated_at = timezone.now()
        header.save(update_fields=["is_locked", "locked_at", "updated_at"])
        messages.success(request, "Form locked.")
        return redirect("requirements:header-detail", pk=pk)


class EditRequestCreateView(View):
    def post(self, request, pk=None, token=None):
        if token:
            header = get_object_or_404(RequirementHeader, public_view_token=token)
        else:
            header = get_object_or_404(RequirementHeader, pk=pk)
        message = request.POST.get("message", "").strip()
        if not message:
            messages.error(request, "Please enter a message describing what needs to be edited.")
            if token:
                return redirect("requirements:public-detail", token=token)
            return redirect("requirements:header-detail", pk=pk)
        EditRequest.objects.create(header=header, event=header.event, message=message)
        messages.success(request, "Edit request sent. Admin will update in 24 hours. NO need of further follow-up")
        if token:
            return redirect("requirements:public-detail", token=token)
        return redirect("requirements:header-detail", pk=pk)


class EditRequestListView(LoginRequiredMixin, View):
    def get(self, request):
        event_id = request.GET.get("event")
        event = None
        if event_id:
            try:
                event = Event.objects.filter(pk=int(event_id), is_active=True).first()
            except (ValueError, TypeError):
                pass
        if not event:
            event = Event.objects.filter(is_current=True, is_active=True).first()
        if not event:
            messages.error(request, "No active event found.")
            return redirect("dashboard:home")
        edit_requests = list(
            EditRequest.objects.filter(event=event)
            .select_related("header", "resolved_by")
            .order_by("-created_at")
        )
        return render(request, "requirements/edit_request_list.html", {
            "edit_requests": edit_requests,
            "event": event,
            "lang": _lang_code(request),
        })


class EditRequestResolveView(LoginRequiredMixin, View):
    def post(self, request, pk):
        er = get_object_or_404(EditRequest, pk=pk)
        resolution = request.POST.get("resolution", "")
        if resolution not in ("SOLVED", "REJECTED"):
            messages.error(request, "Invalid resolution.")
            return redirect("requirements:edit-request-list")
        er.is_resolved = True
        er.resolution = resolution
        er.resolved_by = request.user
        er.resolved_at = timezone.now()
        er.save(update_fields=["is_resolved", "resolution", "resolved_by", "resolved_at"])
        messages.success(request, f"Edit request marked as {resolution.lower()}.")
        return redirect("requirements:edit-request-list")


class RequirementCollectByEventView(RequirementCollectionView):
    pass


class RequirementHeaderCreateView(EventScopedCreateView):
    model = RequirementHeader
    form_class = RequirementHeaderForm
    template_name = "common/form.html"
    success_url = reverse_lazy("requirements:header-list")

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["page_title"] = "Create Requirement"
        context["list_url"] = self.success_url
        return context


class RequirementHeaderUpdateView(EventScopedUpdateView):
    model = RequirementHeader
    form_class = RequirementHeaderForm
    template_name = "common/form.html"
    success_url = reverse_lazy("requirements:header-list")

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["page_title"] = "Update Requirement"
        context["list_url"] = self.success_url
        return context


class RequirementHeaderDeleteView(EventScopedDeleteView):
    model = RequirementHeader
    template_name = "common/confirm_delete.html"
    success_url = reverse_lazy("requirements:header-list")

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["list_url"] = self.success_url
        return context


class RequirementLineListView(EventScopedListView):
    model = RequirementLine
    template_name = "common/list.html"
    row_fields = ("requirement", "item", "required_qty", "remarks")
    headers = ["Requirement", "Item", "Required Qty", "Remarks"]
    search_fields = ["requirement__upashray__name", "item__item_name", "remarks"]
    create_url_name = "requirements:line-create"
    edit_url_name = "requirements:line-update"
    delete_url_name = "requirements:line-delete"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["page_title"] = "Requirement Lines"
        context["create_url"] = reverse_lazy(self.create_url_name)
        return context


class RequirementLineCreateView(EventScopedCreateView):
    model = RequirementLine
    form_class = RequirementLineForm
    template_name = "common/form.html"
    success_url = reverse_lazy("requirements:line-list")

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["page_title"] = "Create Requirement Line"
        context["list_url"] = self.success_url
        return context


class RequirementLineUpdateView(EventScopedUpdateView):
    model = RequirementLine
    form_class = RequirementLineForm
    template_name = "common/form.html"
    success_url = reverse_lazy("requirements:line-list")

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["page_title"] = "Update Requirement Line"
        context["list_url"] = self.success_url
        return context


class RequirementLineDeleteView(EventScopedDeleteView):
    model = RequirementLine
    template_name = "common/confirm_delete.html"
    success_url = reverse_lazy("requirements:line-list")

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["list_url"] = self.success_url
        return context


class SpecialRequirementListView(EventScopedListView):
    model = SpecialRequirement
    template_name = "common/list.html"
    row_fields = ("upashray", "get_priority_display", "get_status_display", "description")
    headers = ["Upashray", "Priority", "Status", "Description"]
    search_fields = ["upashray__name", "description"]
    create_url_name = "requirements:special-create"
    edit_url_name = "requirements:special-update"
    delete_url_name = "requirements:special-delete"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["page_title"] = "Special Requirements"
        context["create_url"] = reverse_lazy(self.create_url_name)
        return context


class SpecialRequirementCreateView(EventScopedCreateView):
    model = SpecialRequirement
    form_class = SpecialRequirementForm
    template_name = "common/form.html"
    success_url = reverse_lazy("requirements:special-list")

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["page_title"] = "Create Special Requirement"
        context["list_url"] = self.success_url
        return context


class SpecialRequirementUpdateView(EventScopedUpdateView):
    model = SpecialRequirement
    form_class = SpecialRequirementForm
    template_name = "common/form.html"
    success_url = reverse_lazy("requirements:special-list")

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["page_title"] = "Update Special Requirement"
        context["list_url"] = self.success_url
        return context


class SpecialRequirementDeleteView(EventScopedDeleteView):
    model = SpecialRequirement
    template_name = "common/confirm_delete.html"
    success_url = reverse_lazy("requirements:special-list")

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["list_url"] = self.success_url
        return context


class PublicRequirementListView(View):
    template_name = "public/requests.html"

    def _get_event(self, token=None):
        if token:
            return Event.objects.filter(public_form_token=token, is_active=True).first()
        return Event.objects.filter(is_active=True).order_by("-is_current", "-start_date", "name").first()

    def _get_rows(self, event):
        return list(
            RequirementHeader.objects.filter(event=event)
            .select_related("upashray")
            .prefetch_related("lines__item")
            .exclude(status=RequirementStatus.DRAFT)
            .order_by("-updated_at", "-created_at")
        )

    def _status_summary(self, headers):
        summary = {key: 0 for key, _label in PUBLIC_STATUS_CHOICES}
        for header in headers:
            summary[header.status] = summary.get(header.status, 0) + 1
        return summary

    def get(self, request, token=None):
        event = self._get_event(token)
        headers = self._get_rows(event) if event else []
        total_items = sum(h.lines.count() for h in headers) if headers else 0
        return render(
            request,
            self.template_name,
            {
                "event": event,
                "headers": headers,
                "status_choices": PUBLIC_STATUS_CHOICES,
                "status_summary": self._status_summary(headers),
                "total_items": total_items,
                "public_collect_url": reverse("requirements:public-collect", kwargs={"event_token": event.public_form_token}) if event else None,
                "public_landing_url": reverse("public-landing"),
                "public_items": build_public_item_preview(event) if event else [],
                "requests_count": len(headers),
            },
        )

    @transaction.atomic
    def post(self, request, token=None):
        event = self._get_event(token)
        if event is None:
            messages.error(request, "No active event found.")
            return redirect(reverse_lazy("public-landing"))

        order_token = request.POST.get("order_token")
        new_status = request.POST.get("status")
        header = RequirementHeader.objects.filter(event=event, public_view_token=order_token).first()
        if header is None:
            messages.error(request, "Requirement order not found.")
            return redirect(reverse("public-requests"))

        valid_statuses = {value for value, _label in PUBLIC_STATUS_CHOICES}
        if new_status not in valid_statuses:
            messages.error(request, "Invalid status.")
            return redirect(reverse("public-requests"))

        previous_status = header.status
        header.status = new_status
        header.save(update_fields=["status", "updated_at"])
        if new_status == RequirementStatus.IN_PROGRESS and header.packing_stock_applied_at is None:
            apply_requirement_packing(header)
        if previous_status != new_status:
            messages.success(request, "Status saved.")
        return redirect(reverse("public-requests"))

