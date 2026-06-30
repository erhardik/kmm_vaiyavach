from io import BytesIO

from django.conf import settings
from django.core.mail import EmailMessage
from django.core.management.base import BaseCommand
from django.db.models import Prefetch
from django.utils import timezone

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from apps.masters.models import Event, Item
from apps.requirements.models import RequirementHeader, RequirementLine, RequirementStatus


class Command(BaseCommand):
    help = "Generate and email the response sheet Excel to FORM_BACKUP_EMAIL daily."

    def handle(self, *args, **options):
        to_email = getattr(settings, "FORM_BACKUP_EMAIL", "")
        if not to_email:
            self.stdout.write(self.style.WARNING("FORM_BACKUP_EMAIL not set. Skipping."))
            return
        event = Event.objects.filter(is_current=True, is_active=True).first()
        if not event:
            event = Event.objects.filter(is_active=True).order_by("-start_date").first()
        if not event:
            self.stdout.write(self.style.WARNING("No active event found. Skipping."))
            return

        headers = list(
            RequirementHeader.objects.filter(event=event)
            .exclude(status=RequirementStatus.DRAFT)
            .select_related("upashray")
            .prefetch_related(Prefetch("lines", queryset=RequirementLine.objects.select_related("item")))
            .order_by("created_at")
        )

        all_items = list(
            Item.objects.filter(
                requirement_lines__requirement__event=event,
                requirement_lines__requirement__is_active=True,
            )
            .exclude(requirement_lines__requirement__status=RequirementStatus.DRAFT)
            .distinct()
            .order_by("standard_serial", "pk")
        )

        workbook = Workbook()
        ws = workbook.active
        ws.title = "Response Sheet"

        header_fill = PatternFill("solid", fgColor="DCE9F5")
        total_fill = PatternFill("solid", fgColor="FFF3BF")
        bold_font = Font(bold=True, color="14324F")

        item_col_map = {}
        for idx, item in enumerate(all_items, 1):
            display = item.item_name_gu or item.item_name
            item_col_map[item.pk] = {"col_idx": idx, "header": f"{idx}-{display}"}

        basic_headers = [
            "Sr. No.", "Timestamp", "Form No.", "Order ID", "Requirement Date",
            "Route", "Sub Route", "Pujya Shri", "Thana", "Area",
            "Current Address", "Chaturmas Address", "Chaturmas Entry Date",
            "Volunteer Name", "Volunteer Mobile", "Stay Type",
            "Care Taker Name", "Care Taker Mobile", "Status",
        ]
        all_headers = basic_headers + [v["header"] for v in item_col_map.values()] + ["Total Qty"]
        ws.append(all_headers)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = bold_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        form_count = 0
        totals = [0] * (len(item_col_map) + 1)

        for header in headers:
            form_count += 1
            row_qty_total = 0
            item_qty_map = {}
            for line in header.lines.all():
                qty = float(line.required_qty)
                item_qty_map[line.item_id] = qty
                row_qty_total += qty

            row_data = [
                form_count,
                timezone.localtime(header.created_at).strftime("%d-%b-%Y %H:%M") if header.created_at else "",
                header.form_number or "",
                header.order_number or "",
                header.requirement_date.strftime("%d-%b-%Y") if header.requirement_date else "",
                header.get_route_area_display() or "",
                header.get_route_sub_area_display() or "",
                header.pujya_shri_name or "",
                header.thana_count or "",
                header.area or "",
                header.current_address or "",
                header.chaturmas_place_address or "",
                header.chaturmas_entry_date.strftime("%d-%b-%Y") if header.chaturmas_entry_date else "",
                header.volunteer_name or "",
                header.volunteer_mobile or "",
                header.get_stay_type_display() or "",
                header.caretaker_name or "",
                header.caretaker_mobile or "",
                header.get_status_display(),
            ]
            item_cells = []
            for item_pk, info in item_col_map.items():
                qty = item_qty_map.get(item_pk, 0)
                item_cells.append(qty)
                totals[info["col_idx"] - 1] += qty
            totals[-1] += row_qty_total
            row_data.extend(item_cells)
            row_data.append(row_qty_total)
            ws.append(row_data)

        total_row = ["TOTAL", "", f"{form_count} Forms"] + [""] * (len(basic_headers) - 3)
        for item_pk, info in item_col_map.items():
            total_row.append(totals[info["col_idx"] - 1])
        total_row.append(totals[-1])
        ws.append(total_row)
        for cell in ws[ws.max_row]:
            cell.fill = total_fill
            cell.font = bold_font

        ws.freeze_panes = "A2"
        basic_count = len(basic_headers)
        for col_idx, column_cells in enumerate(ws.columns, 1):
            max_len = 0
            col_letter = column_cells[0].column_letter
            for cell in column_cells:
                try:
                    text = str(cell.value or "")
                except Exception:
                    text = ""
                max_len = max(max_len, len(text))
            ws.column_dimensions[col_letter].width = min(max_len + 2, 50 if col_idx > basic_count else 40)
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row - 1):
            for cell in row:
                if cell.column <= basic_count and cell.column not in (10, 11):
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                elif cell.column > basic_count:
                    cell.alignment = Alignment(horizontal="right", vertical="center")

        buffer = BytesIO()
        workbook.save(buffer)

        subject = f"Response Sheet - {event.name} - {timezone.localdate().strftime('%d-%b-%Y')}"
        body = f"Daily response sheet for {event.name}.\n\nTotal forms: {form_count}\nGenerated at: {timezone.localtime(timezone.now()).strftime('%d-%b-%Y %H:%M')} IST"
        email = EmailMessage(
            subject=subject,
            body=body,
            from_email=settings.DEFAULT_FROM_EMAIL,
            to=[to_email],
        )
        email.attach(f"response_sheet_{event.name}_{timezone.localdate().strftime('%Y%m%d')}.xlsx", buffer.getvalue(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        email.send()
        self.stdout.write(self.style.SUCCESS(f"Response sheet sent to {to_email}"))
